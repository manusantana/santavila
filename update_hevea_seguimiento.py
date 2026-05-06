#!/usr/bin/env python3
"""
update_hevea_seguimiento.py

Genera (o regenera) dos hojas en Santavila.xlsx con el seguimiento de tarifas
de Hevea, leyendo todos los CSVs con prefijo YYYYMMDD en proveedores_raw/hevea/.

Idempotente: ejecutar varias veces produce el mismo resultado.
NO toca las hojas existentes "Todos", "Hevea", "Balliu".

Hojas que crea/regenera:
  - "Hevea Histórico"   (formato long: una fila por SKU+Producto+Fecha)
  - "Hevea Seguimiento" (formato wide: una fila por SKU+Producto, KPIs actuales)

Aviso: si añades columnas custom a estas dos hojas, se perderán al regenerar.

Requiere openpyxl: pip install openpyxl
"""

import csv
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = Path(__file__).parent
XLSX = BASE / "Santavila.xlsx"
HEVEA_DIR = BASE / "proveedores_raw" / "hevea"

DATE_PREFIX_RE = re.compile(r"^(\d{4})(\d{2})(\d{2})\s*-")

# ── Lectura de CSVs ───────────────────────────────────────────────────────────

def find_dated_csvs(folder):
    out = []
    skipped = []
    for p in folder.glob("*.csv"):
        m = DATE_PREFIX_RE.match(p.name)
        if not m:
            continue
        # Validar que la fecha sea real (rechazar 99/99, 13/45, 02/30, etc.)
        try:
            datetime.strptime(f"{m.group(1)}{m.group(2)}{m.group(3)}", "%Y%m%d")
        except ValueError:
            skipped.append(p.name)
            continue
        date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        out.append((date_str, p))
    if skipped:
        print(f"⚠ CSVs con fecha inválida en el nombre, ignorados: {skipped}", file=sys.stderr)
    return sorted(out)

def normalize_price_col(headers):
    for h in headers:
        if h and "exworks" in h.lower():
            return h
    raise KeyError(f"No se encontró columna 'exworks' en headers: {list(headers)}")

def read_csv(path):
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return []
    pc = normalize_price_col(rows[0].keys())
    out = []
    for r in rows:
        sku = (r.get("SKU") or "").strip()
        prod = (r.get("Producto") or "").strip()
        if not sku:
            continue
        try:
            precio = float((r.get(pc) or "").strip()) if (r.get(pc) or "").strip() else None
        except ValueError:
            precio = None
        try:
            pvp_raw = (r.get("PVP Recomendado") or "").strip()
            pvp = float(pvp_raw) if pvp_raw else None
        except ValueError:
            pvp = None
        out.append({"sku": sku, "producto": prod, "precio": precio, "pvp": pvp})
    return out

# ── Cruce con Shopify Handle desde la hoja "Hevea" existente ─────────────────

def load_sku_to_handle(wb):
    """Mapa (sku, first_token(producto)) -> handle desde la hoja 'Hevea'."""
    if "Hevea" not in wb.sheetnames:
        return {}
    ws = wb["Hevea"]
    mapping = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[2]:
            continue
        sku = str(row[2]).strip()
        prod = str(row[3]).strip() if row[3] else ""
        handle = (row[1] or "").strip() if row[1] else ""
        mapping[(sku, first_token(prod))] = handle
    return mapping

# ── Tendencia con bloques unicode ────────────────────────────────────────────

BLOCKS = "▁▂▃▄▅▆▇█"

def sparkline_text(values):
    nums = [v for v in values if v is not None]
    if len(nums) < 2:
        return ""
    mn, mx = min(nums), max(nums)
    if mx == mn:
        return BLOCKS[3] * len(values)
    rng = mx - mn
    out = ""
    for v in values:
        if v is None:
            out += " "
        else:
            idx = int((v - mn) / rng * (len(BLOCKS) - 1))
            out += BLOCKS[idx]
    return out

# ── Construcción del modelo ──────────────────────────────────────────────────

def first_token(s):
    """
    Primera palabra/token alfanumérico del Producto, normalizado a mayúsculas sin tildes.
    Robusto frente a caracteres no-ASCII al inicio (Ø, º, emojis, etc.): los salta y
    busca el primer token que empiece por letra o dígito ASCII. Si tras la búsqueda
    no encuentra nada útil, devuelve un fallback con los primeros 8 caracteres
    crudos en mayúsculas — mejor un token aproximado que una clave vacía que
    fusionaría productos distintos.
    """
    if not s:
        return ""
    n = unicodedata.normalize("NFKD", s)
    n = "".join(c for c in n if not unicodedata.combining(c))
    m = re.search(r"[A-Za-z0-9][A-Za-z0-9.\-]*", n)
    if m:
        return m.group(0).upper()
    # Fallback: ningún token ASCII alfanumérico — usamos los primeros caracteres
    # del original normalizado para no producir clave vacía.
    fallback = re.sub(r"\s+", "", n)[:8].upper()
    return fallback or "_UNKNOWN_"

def detect_real_duplicated_skus(snapshots):
    """
    SKUs realmente duplicados por el proveedor: aquellos que aparecen >1 vez
    en el mismo CSV (no entre CSVs distintos — eso es solo cambio de nombre).
    """
    real_dups = set()
    for _, rows in snapshots:
        counts = defaultdict(int)
        for r in rows:
            counts[r["sku"]] += 1
        for sku, c in counts.items():
            if c > 1:
                real_dups.add(sku)
    return real_dups

def build_data(snapshots):
    """
    Devuelve:
      products: dict de clave -> {sku, producto, fechas: {fecha: {precio, pvp, producto}}}
      duplicated_skus: set de SKUs duplicados dentro del mismo CSV (flag ⚠)
      reassigned_skus: set de SKUs reasignados entre CSVs distintos (flag ⤺)

    Clave: SIEMPRE (sku, first_token(producto)).
    Esto une acortamientos de nombre (mismo first_token) y separa reasignaciones
    reales (first_token distinto), evitando deltas absurdos cuando el proveedor
    reusa un SKU para un producto totalmente distinto en otro snapshot.

    El campo `producto` guarda el nombre del snapshot más reciente.
    """
    real_dups = detect_real_duplicated_skus(snapshots)

    def make_key(sku, producto):
        return (sku, first_token(producto))

    products = defaultdict(lambda: {"sku": None, "producto": None, "fechas": {}})
    for fecha, rows in sorted(snapshots, key=lambda x: x[0]):
        for r in rows:
            key = make_key(r["sku"], r["producto"])
            products[key]["sku"] = r["sku"]
            products[key]["producto"] = r["producto"]
            products[key]["fechas"][fecha] = {
                "precio": r["precio"],
                "pvp": r["pvp"],
                "producto": r["producto"],
            }

    # SKUs reasignados: aparecen con >1 first_token a lo largo del histórico
    sku_tokens = defaultdict(set)
    for (sku, tok) in products:
        sku_tokens[sku].add(tok)
    reassigned = {sku for sku, toks in sku_tokens.items() if len(toks) > 1}

    return products, real_dups, reassigned

# ── Estilos compartidos ──────────────────────────────────────────────────────

HEVEA_BLUE = "1F4E79"
ROW_A = "DDEEFF"
ROW_B = "EEF5FF"
TOTAL_BG = "F5F5DC"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Hoja "Hevea Histórico" (long) ────────────────────────────────────────────

HEADERS_HIST = [
    "Proveedor", "SKU", "Producto", "Fecha",
    "Precio neto exworks €", "PVP recomendado €",
    "Δ Precio €", "Δ Precio %",
    "Δ PVP €", "Δ PVP %",
]
WIDTHS_HIST = [10, 14, 45, 12, 16, 16, 12, 12, 12, 12]

def write_historico(wb, products):
    if "Hevea Histórico" in wb.sheetnames:
        del wb["Hevea Histórico"]
    ws = wb.create_sheet("Hevea Histórico")

    for ci, h in enumerate(HEADERS_HIST, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEVEA_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ci, w in enumerate(WIDTHS_HIST, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Borde grueso superior para marcar cambio de producto
    thick_top = Side(style="medium", color="1F4E79")
    border_group_top = Border(left=THIN, right=THIN, top=thick_top, bottom=THIN)

    sorted_keys = sorted(products.keys(), key=lambda k: (k[0], k[1] or ""))
    rn = 2
    group_idx = 0  # alterna por producto, no por fila
    for key in sorted_keys:
        sku = key[0]
        info = products[key]
        prev_precio = prev_pvp = None
        group_idx += 1
        # Color del grupo: misma tonalidad para todas las filas del producto
        group_color = ROW_A if group_idx % 2 == 0 else ROW_B
        fechas_ord = sorted(info["fechas"].keys())
        for i_fecha, fecha in enumerate(fechas_ord):
            d = info["fechas"][fecha]
            precio, pvp = d["precio"], d["pvp"]
            producto_en_fecha = d.get("producto", info["producto"])

            d_precio = (precio - prev_precio) if (precio is not None and prev_precio is not None) else None
            d_precio_pct = (d_precio / prev_precio * 100) if (d_precio is not None and prev_precio) else None
            d_pvp = (pvp - prev_pvp) if (pvp is not None and prev_pvp is not None) else None
            d_pvp_pct = (d_pvp / prev_pvp * 100) if (d_pvp is not None and prev_pvp) else None

            row_vals = [
                "Hevea", sku, producto_en_fecha, fecha,
                precio, pvp,
                d_precio, d_precio_pct,
                d_pvp, d_pvp_pct,
            ]
            fill = PatternFill("solid", fgColor=group_color)
            # Solo la primera fila del grupo lleva borde superior grueso
            row_border = border_group_top if i_fecha == 0 else BORDER
            es_primera_del_grupo = (i_fecha == 0)
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=rn, column=ci, value=val)
                c.fill = fill
                c.border = row_border
                # SKU/Producto siempre rellenos (filtros), pero la primera fila del
                # grupo va en negrita azul para que el ojo perciba el cambio.
                if ci in (1, 2, 3) and es_primera_del_grupo:
                    c.font = Font(size=9, bold=True, color="1F4E79")
                else:
                    c.font = Font(size=9)
                if ci in (5, 6):
                    c.number_format = '€ #,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif ci in (7, 9):
                    c.number_format = '+€ #,##0.00;-€ #,##0.00;"—"'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif ci in (8, 10):
                    c.number_format = '+0.0"%";-0.0"%";"—"'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif ci == 4:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            rn += 1
            prev_precio, prev_pvp = precio, pvp

    last_row = rn - 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_HIST))}{last_row}"
    return last_row - 1

# ── Hoja "Hevea Seguimiento" (wide) ──────────────────────────────────────────

HEADERS_SEG = [
    "Proveedor", "Handle Shopify", "SKU", "Producto",
    "1ª aparición", "Última aparición", "Estado",
    "Precio actual €", "PVP actual €", "Margen €",
    "Margen %\nsobre PVP", "Markup %\nsobre coste",
    "Δ Precio %\nvs anterior", "Δ Precio %\nvs origen",
    "Nº subidas", "Tendencia",
]
WIDTHS_SEG = [10, 38, 14, 42, 13, 13, 17, 14, 14, 12, 13, 13, 14, 14, 11, 14]

def write_seguimiento(wb, snapshots, products, duplicated_skus, reassigned_skus, sku_to_handle):
    if "Hevea Seguimiento" in wb.sheetnames:
        del wb["Hevea Seguimiento"]
    ws = wb.create_sheet("Hevea Seguimiento")

    fechas = [f for f, _ in snapshots]
    primera_fecha = fechas[0] if fechas else "—"
    ultima_fecha = fechas[-1] if fechas else "—"

    # Metadata
    title = ws.cell(row=1, column=1, value="Hevea — Seguimiento de tarifas")
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=HEVEA_BLUE)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS_SEG))
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value=f"Última actualización: {ultima_fecha}  │  Fuente: proveedores_raw/hevea/*.csv").font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS_SEG))

    ws.cell(row=3, column=1, value=f"Fechas en seguimiento: {' · '.join(fechas)}  ({len(fechas)} snapshots)").font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(HEADERS_SEG))

    HEADER_ROW = 5
    for ci, h in enumerate(HEADERS_SEG, 1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEVEA_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 38

    for ci, w in enumerate(WIDTHS_SEG, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    sorted_keys = sorted(products.keys(), key=lambda k: (k[0], k[1] or ""))
    rn = HEADER_ROW + 1
    for key in sorted_keys:
        sku = key[0]
        info = products[key]
        producto_nombre = info["producto"]  # último nombre conocido
        fechas_app = sorted(info["fechas"].keys())
        primera_app, ultima_app = fechas_app[0], fechas_app[-1]

        # Estado base por ciclo de vida del PRODUCTO (clave SKU+first_token).
        if ultima_app != ultima_fecha:
            estado_base = "DESCATALOGADO"
        elif primera_app != primera_fecha:
            estado_base = "NUEVO"
        else:
            estado_base = "ACTIVO"
        # Flags ortogonales sobre el SKU (se acumulan):
        #   ⚠ = SKU duplicado dentro de un mismo CSV (proveedor lo usa para >1 producto a la vez)
        #   ⤺ = SKU reasignado entre CSVs (proveedor cambió de producto en otro snapshot)
        flags = ""
        if sku in duplicated_skus:
            flags += " ⚠"
        if sku in reassigned_skus:
            flags += " ⤺"
        estado = f"{estado_base}{flags}"

        ult = info["fechas"][ultima_app]
        precio, pvp = ult["precio"], ult["pvp"]

        margen_e = (pvp - precio) if (pvp is not None and precio is not None) else None
        margen_pct = (margen_e / pvp * 100) if (margen_e is not None and pvp) else None
        markup_pct = (margen_e / precio * 100) if (margen_e is not None and precio) else None

        d_pct_anterior = None
        if len(fechas_app) >= 2:
            penult = info["fechas"][fechas_app[-2]]
            if penult["precio"] and precio is not None:
                d_pct_anterior = (precio - penult["precio"]) / penult["precio"] * 100

        d_pct_origen = None
        primero = info["fechas"][primera_app]
        if primero["precio"] and precio is not None and len(fechas_app) >= 2:
            d_pct_origen = (precio - primero["precio"]) / primero["precio"] * 100

        precios_serie = [info["fechas"][f]["precio"] for f in fechas_app]
        n_subidas = sum(
            1 for i in range(1, len(precios_serie))
            if precios_serie[i] is not None and precios_serie[i-1] is not None
            and precios_serie[i] > precios_serie[i-1]
        )
        tendencia = sparkline_text(precios_serie)

        handle = sku_to_handle.get(key, "")

        row_vals = [
            "Hevea", handle, sku, producto_nombre,
            primera_app, ultima_app, estado,
            precio, pvp, margen_e,
            margen_pct, markup_pct,
            d_pct_anterior, d_pct_origen,
            n_subidas, tendencia,
        ]
        fill = PatternFill("solid", fgColor=ROW_A if rn % 2 == 0 else ROW_B)
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=rn, column=ci, value=val)
            c.fill = fill
            c.border = BORDER
            c.font = Font(size=9)

            if ci in (8, 9, 10):
                c.number_format = '€ #,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (11, 12):
                c.number_format = '0.0"%"'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (13, 14):
                c.number_format = '+0.0"%";-0.0"%";"—"'
                c.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    if val > 0.5:
                        c.font = Font(size=9, color="9C0006", bold=True)
                    elif val < -0.5:
                        c.font = Font(size=9, color="2E7D32", bold=True)
            elif ci == 15:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 16:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Menlo", size=12, color="1F4E79")
            elif ci in (5, 6):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 7:
                c.alignment = Alignment(horizontal="center", vertical="center")
                # Color por ciclo de vida; los flags ⚠/⤺ pasan a rojo negrita
                if estado_base == "NUEVO":
                    c.font = Font(size=9, color="2E7D32", bold=True)
                elif estado_base == "DESCATALOGADO":
                    c.font = Font(size=9, color="666666", italic=True)
                else:
                    c.font = Font(size=9, color="2E7D32")
                if (sku in duplicated_skus) or (sku in reassigned_skus):
                    c.font = Font(size=9, color="9C0006", bold=True, italic=(estado_base == "DESCATALOGADO"))
            elif ci == 2:
                c.alignment = Alignment(horizontal="left", vertical="center")
                if not val:
                    c.font = Font(size=9, color="9C0006", italic=True)
                    c.value = "(no en Shopify)"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        rn += 1

    last_data = rn - 1

    # Fila TOTAL/MEDIA
    total_row = rn
    label = ws.cell(row=total_row, column=4, value="TOTAL / MEDIA")
    label.font = Font(bold=True, size=10)
    label.fill = PatternFill("solid", fgColor=TOTAL_BG)
    label.alignment = Alignment(horizontal="right", vertical="center")
    label.border = BORDER

    def total_cell(col_idx, formula, fmt):
        c = ws.cell(row=total_row, column=col_idx, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER

    for ci in (8, 9, 10):
        col = get_column_letter(ci)
        total_cell(ci, f"=SUM({col}{HEADER_ROW+1}:{col}{last_data})", '€ #,##0.00')
    for ci in (11, 12):
        col = get_column_letter(ci)
        total_cell(ci, f"=AVERAGE({col}{HEADER_ROW+1}:{col}{last_data})", '0.0"%"')
    for ci in (13, 14):
        col = get_column_letter(ci)
        total_cell(ci, f"=AVERAGE({col}{HEADER_ROW+1}:{col}{last_data})", '+0.0"%";-0.0"%";"—"')
    col_o = get_column_letter(15)
    total_cell(15, f"=SUM({col_o}{HEADER_ROW+1}:{col_o}{last_data})", '0')

    # Filtro automático
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(HEADERS_SEG))}{last_data}"
    # Congelar paneles: a la derecha de Producto, debajo de cabecera
    ws.freeze_panes = "E6"

    # Formato condicional sobre Margen % (col K = 11)
    margen_col = get_column_letter(11)
    ws.conditional_formatting.add(
        f"{margen_col}{HEADER_ROW+1}:{margen_col}{last_data}",
        ColorScaleRule(
            start_type="num", start_value="0",  start_color="F8696B",
            mid_type="num",   mid_value="30",   mid_color="FFEB84",
            end_type="num",   end_value="50",   end_color="63BE7B",
        )
    )

    return last_data - HEADER_ROW

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not XLSX.exists():
        sys.exit(f"No existe {XLSX}")
    if not HEVEA_DIR.exists():
        sys.exit(f"No existe {HEVEA_DIR}")

    print(f"Leyendo CSVs de {HEVEA_DIR.relative_to(BASE)}...")
    csvs = find_dated_csvs(HEVEA_DIR)
    if not csvs:
        sys.exit("No se encontraron CSVs con prefijo YYYYMMDD")

    snapshots = []
    for fecha, path in csvs:
        rows = read_csv(path)
        snapshots.append((fecha, rows))
        print(f"  {fecha}: {len(rows)} filas  ({path.name})")

    products, duplicated_skus, reassigned_skus = build_data(snapshots)
    n_skus_unicos = len({k[0] for k in products})
    print(f"\n{len(products)} productos únicos (SKU+token) sobre {n_skus_unicos} SKUs únicos")
    if duplicated_skus:
        print(f"⚠ SKUs duplicados dentro de un mismo CSV ({len(duplicated_skus)}): {sorted(duplicated_skus)}")
    if reassigned_skus:
        print(f"⤺ SKUs reasignados entre CSVs distintos ({len(reassigned_skus)}): {sorted(reassigned_skus)}")

    print(f"\nAbriendo {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX)
    sku_to_handle = load_sku_to_handle(wb)
    print(f"  Hojas existentes: {wb.sheetnames}")
    print(f"  Handles Shopify mapeados desde hoja 'Hevea': {len(sku_to_handle)}")

    n_hist = write_historico(wb, products)
    print(f"\n✓ 'Hevea Histórico' regenerado — {n_hist} filas")

    n_seg = write_seguimiento(wb, snapshots, products, duplicated_skus, reassigned_skus, sku_to_handle)
    print(f"✓ 'Hevea Seguimiento' regenerado — {n_seg} filas")

    wb.save(XLSX)
    print(f"\n✅ {XLSX.name} actualizado")
    print(f"   Hojas finales: {wb.sheetnames}")

if __name__ == "__main__":
    main()
