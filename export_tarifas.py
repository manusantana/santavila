#!/usr/bin/env python3
"""
Genera tarifas_consolidadas.xlsx con todos los productos de Hevea y Balliu.
Columnas: Proveedor, Handle Shopify, SKU, Producto, Coste (€), Precio Venta (€), Margen (€), Margen (%), PVP Rec. (€), Coste Envío (€)
"""

import csv, io, json, os, urllib.request
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE  = Path(__file__).parent
RAW   = BASE / "proveedores_raw"
TOKEN = os.environ.get("SHOPIFY_ACCESS_TOKEN")
if not TOKEN:
    raise RuntimeError("Set SHOPIFY_ACCESS_TOKEN before running this script.")
SHOP  = "mueblesexterior.myshopify.com"
API   = f"https://{SHOP}/admin/api/2026-01/graphql.json"

# ── Fetch SKU → Shopify handle ─────────────────────────────────────────────────

def fetch_sku_to_handle():
    def gql(query, variables=None):
        payload = json.dumps({"query": query, "variables": variables or {}}).encode()
        req = urllib.request.Request(API, data=payload,
            headers={"X-Shopify-Access-Token": TOKEN, "Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read())["data"]

    mapping, after = {}, None
    while True:
        data = gql("""query($cursor: String) {
          products(first: 100, after: $cursor) {
            edges { node { handle variants(first: 20) {
              edges { node { inventoryItem { sku } } }
            }}}
            pageInfo { hasNextPage endCursor }
          }
        }""", {"cursor": after})
        for e in data["products"]["edges"]:
            p = e["node"]
            for ve in p["variants"]["edges"]:
                sku = ve["node"]["inventoryItem"].get("sku","").strip()
                if sku:
                    mapping[sku] = p["handle"]
        pi = data["products"]["pageInfo"]
        if not pi["hasNextPage"]: break
        after = pi["endCursor"]
    return mapping

# ── Parse price string → float ─────────────────────────────────────────────────

def parse_price(s):
    if not s:
        return None
    s = str(s).replace("\xa0", "").replace(" ", "").replace("€", "").replace("%", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")  # Spanish: 1.234,56 → 1234.56
    # else: already English decimal (e.g. 50.0) — leave as-is
    try:
        return float(s)
    except ValueError:
        return None

# ── Load rows from a provider CSV ─────────────────────────────────────────────

def load_csv(fname, sep, sku_to_handle):
    with open(RAW / fname, encoding="utf-8-sig") as f:
        content = f.read()
    reader = csv.DictReader(io.StringIO(content), delimiter=sep)
    rows = []
    for r in reader:
        sku     = r.get("SKU","").strip()
        if not sku:
            continue
        coste    = parse_price(r.get("Precio neto exworks sin iva",""))
        venta    = parse_price(r.get("Precio Venta",""))
        margen_e = parse_price(r.get("Margen €",""))
        margen_p = parse_price(r.get("Margen%",""))
        pvp      = parse_price(r.get("PVP Recomendado",""))
        envio    = parse_price(r.get("Coste de Envio","") or r.get("",""))

        if coste and venta and margen_e is None:
            margen_e = venta - coste
        if coste and venta and margen_p is None and coste > 0:
            margen_p = (venta - coste) / venta * 100

        rows.append({
            "proveedor": r.get("Proveedor","").strip(),
            "handle":    sku_to_handle.get(sku, ""),
            "sku":       sku,
            "producto":  r.get("Producto","").strip(),
            "coste":     coste,
            "venta":     venta,
            "margen_e":  margen_e,
            "margen_p":  margen_p,
            "pvp":       pvp,
            "envio":     envio if envio else 0.0,
        })
    return rows

# ── Build workbook ─────────────────────────────────────────────────────────────

def euro(v):
    return f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if v is not None else "—"

def pct(v):
    return f"{v:.1f}%".replace(".", ",") if v is not None else "—"

def make_workbook(all_rows):
    wb = openpyxl.Workbook()

    # ── Colors & styles ────────────────────────────────────────────────────────
    COL_HEADER_HEVEA  = "1F4E79"   # dark blue
    COL_HEADER_BALLIU = "375623"   # dark green
    COL_HEADER_TOTAL  = "3A3A3A"   # dark grey
    COL_ROW_HEVEA_A   = "DDEEFF"
    COL_ROW_HEVEA_B   = "EEF5FF"
    COL_ROW_BALLIU_A  = "E2EFDA"
    COL_ROW_BALLIU_B  = "F0F7EB"
    COL_TOTAL_ROW     = "F5F5DC"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def row_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    HEADERS = [
        "Proveedor", "Handle Shopify", "SKU", "Producto",
        "Coste neto\n(sin IVA)", "Precio Venta\n(con IVA 21%)",
        "Margen €", "Margen %", "PVP Recomendado", "Coste Envío"
    ]
    COL_WIDTHS = [12, 45, 42, 52, 18, 20, 14, 12, 18, 14]

    def write_sheet(ws, rows, header_color, fill_a, fill_b, sheet_title):
        ws.title = sheet_title
        ws.freeze_panes = "A2"

        # Header row
        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font       = Font(bold=True, color="FFFFFF", size=10)
            cell.fill       = hdr_fill(header_color)
            cell.alignment  = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            cell.border     = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 32

        # Data rows
        for ri, r in enumerate(rows, 2):
            fill = row_fill(fill_a if ri % 2 == 0 else fill_b)
            vals = [
                r["proveedor"], r["handle"], r["sku"], r["producto"],
                r["coste"], r["venta"],
                r["margen_e"], r["margen_p"],
                r["pvp"], r["envio"] or 0.0,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill   = fill
                cell.border = border
                cell.font   = Font(size=9)
                cell.alignment = Alignment(vertical="center")

                # Format numbers
                # cols: 1=Proveedor 2=Handle 3=SKU 4=Producto 5=Coste 6=Venta 7=Margen€ 8=Margen% 9=PVP 10=Envío
                if ci == 8 and val is not None:   # Margen %
                    cell.number_format = '0.0"%"'
                elif ci in (5, 6, 7, 9, 10) and val is not None:
                    cell.number_format = '€ #,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif ci in (1, 2, 3):
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Color margin % cell
                if ci == 8 and val is not None:
                    if val >= 50:
                        cell.font = Font(size=9, color="1A6B2A", bold=True)
                    elif val >= 35:
                        cell.font = Font(size=9, color="7B6000")
                    else:
                        cell.font = Font(size=9, color="9C0006")

            ws.row_dimensions[ri].height = 16

        # Totals row
        n = len(rows) + 2
        ws.cell(row=n, column=4, value="TOTAL / MEDIA").font = Font(bold=True, size=10)
        ws.cell(row=n, column=4).fill = row_fill(COL_TOTAL_ROW)

        for ci in (5, 6, 7):   # Coste, Venta, Margen€
            col = get_column_letter(ci)
            cell = ws.cell(row=n, column=ci,
                           value=f"=SUM({col}2:{col}{n-1})")
            cell.number_format = '€ #,##0.00'
            cell.font  = Font(bold=True, size=10)
            cell.fill  = row_fill(COL_TOTAL_ROW)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border

        # Avg margin %  (col 8)
        col8 = get_column_letter(8)
        cell = ws.cell(row=n, column=8,
                       value=f"=AVERAGE({col8}2:{col8}{n-1})")
        cell.number_format = '0.0"%"'
        cell.font  = Font(bold=True, size=10)
        cell.fill  = row_fill(COL_TOTAL_ROW)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── Individual sheets ──────────────────────────────────────────────────────
    hevea_rows  = [r for r in all_rows if r["proveedor"] == "Hevea"]
    balliu_rows = [r for r in all_rows if r["proveedor"] == "Balliu"]

    ws_all = wb.active
    write_sheet(ws_all, all_rows,  COL_HEADER_TOTAL,  "F2F2F2", "FAFAFA",  "Todos")

    ws_hevea = wb.create_sheet()
    write_sheet(ws_hevea, hevea_rows, COL_HEADER_HEVEA, COL_ROW_HEVEA_A, COL_ROW_HEVEA_B, "Hevea")

    ws_balliu = wb.create_sheet()
    write_sheet(ws_balliu, balliu_rows, COL_HEADER_BALLIU, COL_ROW_BALLIU_A, COL_ROW_BALLIU_B, "Balliu")

    return wb

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Cargando handles de Shopify...")
    sku_to_handle = fetch_sku_to_handle()
    print(f"  {len(sku_to_handle)} SKUs mapeados")

    hevea_rows  = load_csv("CSV-Catalogo-Santavila.csv", ",", sku_to_handle)
    balliu_rows = load_csv("CSV-Catalogo-Balliu.csv", ";", sku_to_handle)
    all_rows = hevea_rows + balliu_rows

    print(f"Hevea:  {len(hevea_rows)} productos")
    print(f"Balliu: {len(balliu_rows)} productos")
    print(f"Total:  {len(all_rows)} productos")

    wb = make_workbook(all_rows)
    out = BASE / "tarifas_consolidadas.xlsx"
    wb.save(out)
    print(f"\nGuardado: {out}")

if __name__ == "__main__":
    main()
