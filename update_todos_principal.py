#!/usr/bin/env python3
"""
update_todos_principal.py

Actualiza las columnas de datos numéricos (E..I) de la hoja "20260508 -Todos "
de Santavila.xlsx con los precios actuales de los proveedores:

  Hevea: último CSV con prefijo YYYYMMDD en proveedores_raw/hevea/
         (cruce por SKU; productos con SKU duplicado se desambiguan por
          first_token del Producto)

  Balliu: PDF de COSTE más reciente y PDF de PVP RECOMENDADO más reciente
          en proveedores_raw/balliu/ (cruce SKU ↔ PDF construido a partir
          del coste actual de la hoja "Balliu" + reverse-engineering del
          nombre del producto codificado en el SKU largo)

Columnas que actualiza (NO toca A..D, ni K..N que son fórmulas):
  E: Coste neto (sin IVA)
  F: Precio Venta (con IVA 21%)  = PVP Recomendado sin IVA × 1,21
  G: Margen €                    = PVP sin IVA − Coste sin IVA
  H: Margen %                    = Margen € / PVP sin IVA × 100  (margen bruto)
  I: PVP Recomendado             = PVP sin IVA del proveedor

Idempotente: ejecutar varias veces produce el mismo resultado.

Uso:
  python3 update_todos_principal.py

Genera además proveedores_raw/balliu/_sku_mapping.json con el mapeo
construido para Balliu (auditable / reusable).
"""

import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

# Reusar el parser de Hevea y Balliu
sys.path.insert(0, str(Path(__file__).parent))
from update_hevea_seguimiento import find_dated_csvs, read_csv as read_hevea_csv, first_token
from update_balliu_seguimiento import (
    find_dated_pdfs, extract_tarifa as extract_balliu_pdf, assign_ord,
    infer_tipo, TIPO_COSTE, TIPO_PVP, BALLIU_DIR
)

BASE = Path(__file__).parent
XLSX = BASE / "Santavila.xlsx"
TODOS_SHEET = "20260508 -Todos "    # destino del update — ojo: espacio final
GROUND_TRUTH_SHEET = "Todos"        # snapshot intocado, lo usamos para los costes originales
                                    # con los que construimos el mapping fila→PDF (cada fila duplicada
                                    # tiene su coste propio, lo que las desambigua)
BALLIU_SHEET = "Balliu"             # también con costes originales
HEVEA_DIR = BASE / "proveedores_raw" / "hevea"
BALLIU_MAPPING_JSON = BALLIU_DIR / "_sku_mapping.json"
IVA = 1.21


# ── Utilidades ───────────────────────────────────────────────────────────────

def normalize_text(s):
    """Normalizar a ASCII upper sin tildes, sin signos, multi-espacio → single."""
    if not s:
        return ""
    n = unicodedata.normalize("NFKD", s)
    n = "".join(c for c in n if not unicodedata.combining(c))
    n = re.sub(r"[^A-Za-z0-9]+", " ", n)
    return re.sub(r"\s+", " ", n).strip().upper()


def calc_metrics(coste, pvp_sin_iva):
    """Devuelve (precio_venta_con_iva, margen_e, margen_pct, pvp_rec_sin_iva)."""
    if coste is None or pvp_sin_iva is None:
        return None, None, None, pvp_sin_iva
    precio_venta = round(pvp_sin_iva * IVA, 2)
    margen_e = round(pvp_sin_iva - coste, 2)
    margen_pct = round(margen_e / pvp_sin_iva * 100, 2) if pvp_sin_iva else None
    return precio_venta, margen_e, margen_pct, pvp_sin_iva


# ── Hevea: snapshot actual ───────────────────────────────────────────────────

def load_hevea_actual():
    """Devuelve dict (sku, first_token(producto)) -> {coste, pvp} del CSV más reciente."""
    csvs = find_dated_csvs(HEVEA_DIR)
    if not csvs:
        sys.exit("No hay CSVs de Hevea")
    fecha, path = csvs[-1]
    rows = read_hevea_csv(path)
    out = {}
    # Si el SKU es único, mapeamos por SKU directo. Si está duplicado en el mismo CSV,
    # mapeamos por (sku, first_token) — el script de Hevea hace lo mismo.
    by_sku = defaultdict(list)
    for r in rows:
        by_sku[r["sku"]].append(r)
    for sku, items in by_sku.items():
        if len(items) == 1:
            r = items[0]
            out[(sku, None)] = {"coste": r["precio"], "pvp": r["pvp"], "producto": r["producto"]}
            # también accesible por first_token para los SKUs reusados que el Todos
            # ya tiene como dos filas distintas (mismo SKU, distinto Producto)
            out[(sku, first_token(r["producto"]))] = out[(sku, None)]
        else:
            # SKU duplicado dentro del mismo CSV (557-010147, 557-010884, 557-1563)
            for r in items:
                out[(sku, first_token(r["producto"]))] = {
                    "coste": r["precio"], "pvp": r["pvp"], "producto": r["producto"]
                }
    return fecha, out


# ── Balliu: mapping SKU ↔ (P, V, G, ord) y datos actuales ───────────────────

def extract_sku_token(sku):
    """Del SKU 'BALLIU_EVA_PRO_TUMBONA_CHASIS_BLANCO_TA_923110D9' extrae
    'EVA PRO TUMBONA CHASIS BLANCO TA' (sin prefijo BALLIU_, sin hash final).
    Útil para desambiguar entre matches con mismo coste."""
    if not sku:
        return ""
    s = sku
    if s.startswith("BALLIU_"):
        s = s[len("BALLIU_"):]
    parts = s.split("_")
    # quitar hash al final si tiene 8 chars hex
    while parts and re.fullmatch(r"[A-F0-9]{6,12}", parts[-1]):
        parts.pop()
    # quitar tokens vacíos (puede haber doble guion bajo)
    parts = [p for p in parts if p]
    return " ".join(parts).upper()


def build_balliu_row_mapping(wb):
    """Construye un mapping POR FILA (no por SKU global): cada fila Balliu
    de la hoja "Todos" antigua se cruza con el PDF 30/03 usando su coste
    original como clave principal.

    Razón: hay 5 SKUs Balliu que aparecen en ≥2 filas distintas con costes
    distintos (datos históricos donde el SKU está mal etiquetado o se reusó).
    Mapear por SKU global pierde esa diferenciación. La fila + su coste
    propio sí identifica la combinación correcta del PDF.

    Devuelve {(sku, fila_todos): {"producto","variante","grupo","ord"}}.
    """
    pdfs = find_dated_pdfs(BALLIU_DIR)
    pdf_costes = [(f, p) for f, p in pdfs if infer_tipo(p.name) == TIPO_COSTE]
    if not pdf_costes:
        sys.exit("No hay PDF de COSTE en proveedores_raw/balliu/")
    fecha_coste, path_coste = pdf_costes[-1]
    pdf_rows = assign_ord(extract_balliu_pdf(path_coste))

    if GROUND_TRUTH_SHEET not in wb.sheetnames:
        sys.exit(f"No existe la hoja '{GROUND_TRUTH_SHEET}' (snapshot original)")
    ws = wb[GROUND_TRUTH_SHEET]
    todos_balliu_rows = []
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value != "Balliu":
            continue
        sku = ws.cell(r, 3).value
        coste = ws.cell(r, 5).value
        if sku and isinstance(coste, (int, float)):
            todos_balliu_rows.append({"sku": sku, "coste": float(coste), "fila": r})

    pdf_by_coste = defaultdict(list)
    for r in pdf_rows:
        if r["precio"] is not None:
            pdf_by_coste[round(r["precio"], 2)].append(r)

    mapping = {}        # (sku, fila) -> {producto, variante, grupo, ord}
    no_match = []
    used_pdf_keys = set()

    for sr in todos_balliu_rows:
        sku, fila = sr["sku"], sr["fila"]
        coste = round(sr["coste"], 2)
        candidates = pdf_by_coste.get(coste, [])
        if not candidates:
            no_match.append((sku, fila, coste, "coste no encontrado en PDF"))
            continue
        if len(candidates) == 1:
            chosen = candidates[0]
        else:
            sku_norm = normalize_text(extract_sku_token(sku))
            scored = []
            for c in candidates:
                prod_norm = normalize_text(c["producto"])
                words = prod_norm.split()
                score = sum(1 for w in words if w in sku_norm) / len(words) if words else 0
                # bonus si la variante también coincide en el SKU
                var_norm = normalize_text(c["variante"])
                var_words = var_norm.split()
                if var_words:
                    score += 0.3 * sum(1 for w in var_words if w in sku_norm) / len(var_words)
                # penalizar si esa combinación ya está usada
                k = (c["producto"], c["variante"], c["grupo"], c["ord"])
                if k in used_pdf_keys:
                    score -= 1.0
                scored.append((score, c))
            scored.sort(key=lambda x: -x[0])
            chosen = scored[0][1]
        key = (chosen["producto"], chosen["variante"], chosen["grupo"], chosen["ord"])
        if key in used_pdf_keys:
            alt = [c for c in candidates
                   if (c["producto"], c["variante"], c["grupo"], c["ord"]) not in used_pdf_keys]
            if not alt:
                no_match.append((sku, fila, coste, "todas las opciones del PDF ya usadas"))
                continue
            chosen = alt[0]
            key = (chosen["producto"], chosen["variante"], chosen["grupo"], chosen["ord"])
        used_pdf_keys.add(key)
        mapping[(sku, fila)] = {
            "producto": chosen["producto"],
            "variante": chosen["variante"],
            "grupo": chosen["grupo"],
            "ord": chosen["ord"],
        }

    return fecha_coste, mapping, no_match


def load_balliu_actual(wb):
    """Devuelve (fecha_coste, fecha_pvp, datos_por_(sku,fila){...})."""
    pdfs = find_dated_pdfs(BALLIU_DIR)
    snapshots_coste = sorted([(f, p) for f, p in pdfs if infer_tipo(p.name) == TIPO_COSTE])
    snapshots_pvp = sorted([(f, p) for f, p in pdfs if infer_tipo(p.name) == TIPO_PVP])

    fecha_coste = snapshots_coste[-1][0] if snapshots_coste else None
    fecha_pvp = snapshots_pvp[-1][0] if snapshots_pvp else None

    pdf_coste = assign_ord(extract_balliu_pdf(snapshots_coste[-1][1])) if snapshots_coste else []
    pdf_pvp = assign_ord(extract_balliu_pdf(snapshots_pvp[-1][1])) if snapshots_pvp else []

    by_key_coste = {(r["producto"], r["variante"], r["grupo"], r["ord"]): r["precio"] for r in pdf_coste}
    by_key_pvp = {(r["producto"], r["variante"], r["grupo"], r["ord"]): r["precio"] for r in pdf_pvp}

    fecha_coste_meta, mapping, no_match = build_balliu_row_mapping(wb)

    out = {}  # (sku, fila) -> {coste, pvp, key}
    for (sku, fila), k in mapping.items():
        key = (k["producto"], k["variante"], k["grupo"], k["ord"])
        out[(sku, fila)] = {
            "coste": by_key_coste.get(key),
            "pvp": by_key_pvp.get(key),
            "key": k,
        }

    # Persistir mapping (auditable)
    out_json = {
        "_meta": {
            "fecha_coste_origen": fecha_coste,
            "fecha_pvp_origen": fecha_pvp,
            "total_filas_mapeadas": len(mapping),
            "no_match": [{"sku": s, "fila_hoja": f, "coste": c, "razon": r} for s, f, c, r in no_match],
        },
        "mapping": [
            {"sku": sku, "fila_todos_origen": fila, **k}
            for (sku, fila), k in mapping.items()
        ],
    }
    BALLIU_MAPPING_JSON.write_text(json.dumps(out_json, indent=2, ensure_ascii=False))

    return fecha_coste, fecha_pvp, out, no_match


# ── Actualización de la hoja Todos ───────────────────────────────────────────

def update_todos(wb, hevea_actual, hevea_fecha, balliu_actual, balliu_fc, balliu_fp):
    if TODOS_SHEET not in wb.sheetnames:
        sys.exit(f"No existe la hoja '{TODOS_SHEET}' en {XLSX}")
    ws = wb[TODOS_SHEET]

    stats = {"hevea_ok": 0, "hevea_no_match": [], "balliu_ok": 0, "balliu_no_match": [],
             "otros": []}

    for r in range(3, ws.max_row + 1):
        prov = ws.cell(r, 1).value
        sku = ws.cell(r, 3).value
        prod = ws.cell(r, 4).value
        if not prov:
            continue

        if prov == "Hevea":
            tok = first_token(prod) if prod else None
            data = hevea_actual.get((sku, tok)) or hevea_actual.get((sku, None))
            if not data:
                stats["hevea_no_match"].append((r, sku, prod))
                continue
            coste = data["coste"]
            pvp = data["pvp"]
            precio_venta, margen_e, margen_pct, _ = calc_metrics(coste, pvp)
            ws.cell(r, 5).value = coste
            ws.cell(r, 6).value = precio_venta
            ws.cell(r, 7).value = margen_e
            ws.cell(r, 8).value = margen_pct
            ws.cell(r, 9).value = pvp
            stats["hevea_ok"] += 1

        elif prov == "Balliu":
            # mapping clave: (sku, fila) — las filas de "Todos" y "20260508 -Todos "
            # están alineadas, por lo que la fila r aquí coincide con la fila usada
            # al construir el mapping
            data = balliu_actual.get((sku, r))
            if not data or data["coste"] is None or data["pvp"] is None:
                stats["balliu_no_match"].append((r, sku, prod))
                continue
            coste = data["coste"]
            pvp = data["pvp"]
            precio_venta, margen_e, margen_pct, _ = calc_metrics(coste, pvp)
            ws.cell(r, 5).value = coste
            ws.cell(r, 6).value = precio_venta
            ws.cell(r, 7).value = margen_e
            ws.cell(r, 8).value = margen_pct
            ws.cell(r, 9).value = pvp
            stats["balliu_ok"] += 1
        else:
            stats["otros"].append((r, prov, sku))

    return stats


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not XLSX.exists():
        sys.exit(f"No existe {XLSX}")

    print(f"Cargando snapshot Hevea...")
    hevea_fecha, hevea_actual = load_hevea_actual()
    print(f"  Fecha: {hevea_fecha}  ·  {len(hevea_actual)} entradas (incluye duplicados por first_token)")

    print(f"\nAbriendo {XLSX.name} (modo lectura+escritura)...")
    wb = openpyxl.load_workbook(XLSX)
    print(f"  Hojas: {wb.sheetnames}")

    print(f"\nConstruyendo mapping Balliu SKU ↔ PDF...")
    balliu_fc, balliu_fp, balliu_actual, no_match_balliu = load_balliu_actual(wb)
    print(f"  Fecha COSTE: {balliu_fc}  ·  Fecha PVP: {balliu_fp}")
    print(f"  SKUs mapeados: {len(balliu_actual)} / 165")
    print(f"  Mapping persistido en: {BALLIU_MAPPING_JSON.relative_to(BASE)}")
    if no_match_balliu:
        print(f"  ⚠ {len(no_match_balliu)} SKUs sin match:")
        for sku, fila, coste, razon in no_match_balliu[:5]:
            print(f"    R{fila}  {sku}  coste={coste}  → {razon}")
        if len(no_match_balliu) > 5:
            print(f"    ... y {len(no_match_balliu)-5} más")

    print(f"\nActualizando hoja '{TODOS_SHEET}'...")
    stats = update_todos(wb, hevea_actual, hevea_fecha, balliu_actual, balliu_fc, balliu_fp)
    print(f"  Hevea actualizadas:  {stats['hevea_ok']}")
    if stats["hevea_no_match"]:
        print(f"  ⚠ Hevea sin match en CSV {hevea_fecha}: {len(stats['hevea_no_match'])}")
        for r, sku, prod in stats["hevea_no_match"][:5]:
            print(f"    R{r}  {sku}  {prod}")
        if len(stats["hevea_no_match"]) > 5:
            print(f"    ... y {len(stats['hevea_no_match'])-5} más")
    print(f"  Balliu actualizadas: {stats['balliu_ok']}")
    if stats["balliu_no_match"]:
        print(f"  ⚠ Balliu sin datos completos: {len(stats['balliu_no_match'])}")
        for r, sku, prod in stats["balliu_no_match"][:5]:
            print(f"    R{r}  {sku}  {prod}")
        if len(stats["balliu_no_match"]) > 5:
            print(f"    ... y {len(stats['balliu_no_match'])-5} más")

    wb.save(XLSX)
    print(f"\n✅ {XLSX.name} actualizado")
    print(f"   Hoja regenerada: '{TODOS_SHEET}'")


if __name__ == "__main__":
    main()
