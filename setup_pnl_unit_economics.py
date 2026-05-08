#!/usr/bin/env python3
"""
setup_pnl_unit_economics.py

Crea / regenera 6 hojas financieras dentro de Santavila.xlsx:

  00_SUPUESTOS              → Parámetros editables (única fuente de verdad)
  01_PNL_SANTAVILA          → P&L mensual: Conservador | Base | Optimista + break-even
  02_UNIT_ECONOMICS_SKU     → Por SKU: margen contributivo, CAC máx, ROAS mín, categoría
  03_ESCENARIOS_MARKETING   → Calculadora bidireccional (presupuesto ↔ ROAS)
  04_PRODUCTOS_PRIORIDAD    → Top 30 SKUs por margen contributivo, filtros
  05_DASHBOARD              → KPIs en 1 página

NO toca las hojas existentes:
  20260508 -Todos , Todos, Hevea, Balliu, Hevea/Balliu Histórico/Seguimiento

Todas las fórmulas referencian celdas con NOMBRE (DefinedName) en 00_SUPUESTOS.
Editar un valor en 00_SUPUESTOS recalcula automáticamente todo el modelo.

Idempotente: ejecutar varias veces produce el mismo resultado.
Hace backup automático en .backups/ antes de tocar el archivo.

Uso:
  python3 setup_pnl_unit_economics.py
"""

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).parent
XLSX = BASE / "Santavila.xlsx"
BACKUPS = BASE / ".backups"
TODOS_SHEET = "20260508 -Todos "

# ── Estilos ──────────────────────────────────────────────────────────────────
COLOR_HEADER = "1F4E79"        # azul corporativo (mismo que Hevea)
COLOR_SECTION = "2E75B6"       # azul medio
COLOR_EDITABLE = "FFF2CC"      # amarillo suave para celdas editables
COLOR_CALC = "E2EFDA"          # verde suave para celdas calculadas
COLOR_KPI = "FCE4D6"           # naranja suave para KPIs destacados
COLOR_TOTAL = "F5F5DC"         # beige para totales
COLOR_GOOD = "C6EFCE"          # verde
COLOR_WARN = "FFEB9C"          # amarillo
COLOR_BAD = "FFC7CE"           # rojo

THIN = Side(style="thin", color="CCCCCC")
MED = Side(style="medium", color=COLOR_HEADER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def title_cell(ws, row, col, text, span=10, color=COLOR_HEADER, size=14):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=size)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = 28


def section_cell(ws, row, col, text, span=5):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=COLOR_SECTION)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def header_cell(ws, row, col, text, color=COLOR_HEADER):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def editable_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=COLOR_EDITABLE)
    c.font = Font(size=10, bold=True, color="9C5700")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def calc_cell(ws, row, col, formula, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = PatternFill("solid", fgColor=COLOR_CALC)
    c.font = Font(size=10, bold=bold, color="375623")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def kpi_cell(ws, row, col, formula, fmt=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = PatternFill("solid", fgColor=COLOR_KPI)
    c.font = Font(size=11, bold=True, color="C65911")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = Border(left=MED, right=MED, top=MED, bottom=MED)
    if fmt:
        c.number_format = fmt
    return c


def label_cell(ws, row, col, text, indent=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    c.border = BORDER
    return c


def note_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=9, italic=True, color="666666")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c


# ── DefinedNames helper ──────────────────────────────────────────────────────

def add_named(wb, name, sheet, ref):
    """Define un nombre global para que las fórmulas digan 'AOV' en vez de
    "'00_SUPUESTOS'!$B$28". Mucho más legible y mantenible."""
    dn = DefinedName(name=name, attr_text=f"'{sheet}'!{ref}")
    wb.defined_names[name] = dn


# ── 00_SUPUESTOS ─────────────────────────────────────────────────────────────

# Estructura: cada entrada es (var_name, label, default, unit, fmt, note)
# fmt: '€', '%', 'int', 'ratio', 'text'
SUPUESTOS = [
    ("__SECTION", "PLATAFORMA & COSTES FIJOS", None, None, None, None),
    ("Shopify_plan", "Plan Shopify (mensual)", 29, "€/mes", "€",
     "Plan Basic 29€/mes confirmado por usuario."),
    ("Dominio", "Dominio (prorrateado)", 1, "€/mes", "€",
     "12€/año aprox / 12 meses."),
    ("Apps", "Apps Shopify (B2B, etc.)", 30, "€/mes", "€",
     "Estimación inicial; ajustar cuando se contraten."),
    ("Personal_fijo", "Personal fijo (sueldos)", 0, "€/mes", "€",
     "0€ por ahora — los socios cobran del beneficio. Editar cuando haya nómina."),

    ("__SECTION", "COMISIONES DE PAGO", None, None, None, None),
    ("Comision_pago_pct", "Comisión Shopify Payments (% sobre venta)", 0.015, "%", "%",
     "1,5% medio España. Stripe ~1,4%+0,25; PayPal ~3,4%+0,35."),
    ("Comision_pago_fija", "Comisión fija por transacción", 0.25, "€/pedido", "€",
     "Coste fijo independiente del importe."),

    ("__SECTION", "LOGÍSTICA SALIDA", None, None, None, None),
    ("Hevea_umbral_envio_gratis", "Hevea — umbral envío gratis", 900, "€", "€",
     "Pedidos > umbral van gratis a península (acuerdo con Hevea)."),
    ("Hevea_coste_envio_bajo", "Hevea — coste envío si < umbral", 30, "€/pedido", "€",
     "Estimación; producto voluminoso. Ajustar con tarifa real del transportista."),
    ("Balliu_coste_envio", "Balliu — coste envío medio", 30, "€/pedido", "€",
     "Sin acuerdo confirmado. Estimación misma magnitud que Hevea bajo umbral."),
    ("Coste_medio_envio_pedido", "Coste medio mezcla del catálogo (asignado a pedido)", 22, "€/pedido", "€",
     "Asume mezcla: 50% Hevea (la mitad >900€ gratis) + 50% Balliu siempre con coste."),

    ("__SECTION", "DEVOLUCIONES E INCIDENCIAS", None, None, None, None),
    ("Tasa_devolucion", "Tasa de devolución (% pedidos)", 0.04, "%", "%",
     "4% — típico mobiliario premium (vs 20-30% moda). Ajustar con datos reales."),
    ("Coste_devolucion", "Coste medio por devolución", 50, "€", "€",
     "Logística inversa de producto voluminoso."),
    ("Tasa_incidencia", "Tasa de incidencias transporte (% pedidos)", 0.02, "%", "%",
     "Roturas, daños, retraso indemnizado."),
    ("Coste_incidencia", "Coste medio por incidencia", 80, "€", "€",
     "Reposición parcial, descuentos comerciales, gestión."),

    ("__SECTION", "MARKETING — ROAS OBJETIVO", None, None, None, None),
    ("ROAS_conservador", "ROAS objetivo — escenario CONSERVADOR", 5, "x", "ratio",
     "Más conservador = invertimos menos, beneficio cierto. ROAS 5 = 1€ → 5€ ventas."),
    ("ROAS_base", "ROAS objetivo — escenario BASE", 4, "x", "ratio",
     "Realista para mobiliario premium con paid media bien optimizado."),
    ("ROAS_optimista", "ROAS objetivo — escenario OPTIMISTA", 3, "x", "ratio",
     "Más optimista = más volumen, más riesgo. ROAS 3 = punto de inflexión."),

    ("__SECTION", "WEB & CONVERSIÓN", None, None, None, None),
    ("Tasa_conversion", "Tasa de conversión web (sesión → pedido)", 0.008, "%", "%",
     "0,8% — referencia mobiliario premium. Mejorable con UX y trust."),
    ("AOV", "AOV / ticket medio objetivo (con IVA)", 500, "€", "€",
     "500€ confirmado por usuario."),

    ("__SECTION", "VOLUMEN ESCENARIO BASE", None, None, None, None),
    ("Pedidos_mes_conservador", "Pedidos/mes — CONSERVADOR", 15, "pedidos", "int",
     "Arranque suave, 1 pedido cada 2 días."),
    ("Pedidos_mes_base", "Pedidos/mes — BASE", 30, "pedidos", "int",
     "1 pedido al día. Realista mes 6-12."),
    ("Pedidos_mes_optimista", "Pedidos/mes — OPTIMISTA", 60, "pedidos", "int",
     "2 pedidos al día. Crecimiento sostenido."),

    ("__SECTION", "FISCAL", None, None, None, None),
    ("IVA", "IVA aplicado", 0.21, "%", "%",
     "21% España. Incluido en PVP de la hoja 20260508 -Todos."),

    ("__SECTION", "UMBRALES DE CATEGORIZACIÓN SKU", None, None, None, None),
    ("Umbral_PUSH_pct", "PUSH — Margen contributivo % mínimo", 0.30, "%", "%",
     "≥30% margen contributivo → producto a empujar comercialmente."),
    ("Umbral_PUSH_eur", "PUSH — Margen contributivo € mínimo", 100, "€", "€",
     "≥100€/unidad. Combinado con %."),
    ("Umbral_NEUTRAL_pct", "NEUTRAL — Margen contributivo % mínimo", 0.15, "%", "%",
     "15-30% margen → vender pero sin push agresivo."),
    ("Umbral_WATCH_pct", "WATCH — Margen contributivo % mínimo", 0.05, "%", "%",
     "5-15% margen → vigilar, mejorar coste o subir precio."),
]


def write_supuestos(wb):
    name = "00_SUPUESTOS"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index=0)

    title_cell(ws, 1, 1, "00 · SUPUESTOS — única fuente de verdad del modelo financiero", span=5)
    note_cell(ws, 2, 1, "Edita los valores en amarillo. Todas las hojas (P&L, Unit Economics, Escenarios, Dashboard) recalculan automáticamente.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    # Cabeceras
    HEADER_ROW = 4
    headers = ["Variable", "Etiqueta", "Valor", "Unidad", "Nota / fuente"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, HEADER_ROW, ci, h)
    ws.row_dimensions[HEADER_ROW].height = 28

    # Anchos
    ws.column_dimensions["A"].width = 30  # variable
    ws.column_dimensions["B"].width = 50  # etiqueta
    ws.column_dimensions["C"].width = 16  # valor
    ws.column_dimensions["D"].width = 14  # unidad
    ws.column_dimensions["E"].width = 60  # nota

    rn = HEADER_ROW + 1
    for entry in SUPUESTOS:
        var, label, default, unit, fmt, note = entry
        if var == "__SECTION":
            section_cell(ws, rn, 1, label, span=5)
            ws.row_dimensions[rn].height = 22
            rn += 1
            continue

        ws.cell(row=rn, column=1, value=var).font = Font(name="Menlo", size=9, color="666666")
        ws.cell(row=rn, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=rn, column=1).border = BORDER

        label_cell(ws, rn, 2, label)

        # Valor editable
        if fmt == "€":
            ec = editable_cell(ws, rn, 3, default, '€ #,##0.00')
        elif fmt == "%":
            ec = editable_cell(ws, rn, 3, default, '0.0%')
        elif fmt == "ratio":
            ec = editable_cell(ws, rn, 3, default, '0.0"x"')
        elif fmt == "int":
            ec = editable_cell(ws, rn, 3, default, '0')
        else:
            ec = editable_cell(ws, rn, 3, default)

        # DefinedName apuntando a esta celda
        ref = f"$C${rn}"
        add_named(wb, var, name, ref)

        # Unidad
        u = ws.cell(row=rn, column=4, value=unit)
        u.font = Font(size=9, color="666666")
        u.alignment = Alignment(horizontal="center", vertical="center")
        u.border = BORDER

        note_cell(ws, rn, 5, note)
        ws.cell(row=rn, column=5).border = BORDER

        rn += 1

    # Filas calculadas (totales agregados, también nombrados)
    rn += 1
    section_cell(ws, rn, 1, "AGREGADOS CALCULADOS (no editar)", span=5)
    rn += 1

    label_cell(ws, rn, 2, "Total costes fijos mensuales")
    f = f"=Shopify_plan+Dominio+Apps+Personal_fijo"
    calc_cell(ws, rn, 3, f, '€ #,##0.00', bold=True)
    add_named(wb, "TOTAL_FIJOS_MES", name, f"$C${rn}")
    note_cell(ws, rn, 5, "Suma de plan + dominio + apps + personal.")
    rn += 1

    label_cell(ws, rn, 2, "Coste variable estimado por pedido")
    # Comisión + logística + devoluciones + incidencias (todo asignado a 1 pedido medio)
    f = (f"=AOV*Comision_pago_pct + Comision_pago_fija "
         f"+ Coste_medio_envio_pedido "
         f"+ Tasa_devolucion*Coste_devolucion "
         f"+ Tasa_incidencia*Coste_incidencia")
    calc_cell(ws, rn, 3, f, '€ #,##0.00', bold=True)
    add_named(wb, "VAR_COST_PER_ORDER", name, f"$C${rn}")
    note_cell(ws, rn, 5, "Comisión pago + envío + (% dev × coste dev) + (% incid × coste incid).")
    rn += 1

    ws.freeze_panes = "A5"


# ── 01_PNL_SANTAVILA ─────────────────────────────────────────────────────────

def write_pnl(wb):
    name = "01_PNL_SANTAVILA"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    title_cell(ws, 1, 1, "01 · P&L MENSUAL — escenarios Conservador / Base / Optimista", span=5)
    note_cell(ws, 2, 1, "Modelo mensual operativo. Editar valores solo en 00_SUPUESTOS — esta hoja recalcula sola.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    # Headers
    H = 4
    ws.cell(row=H, column=1, value="Concepto").font = Font(bold=True, color="FFFFFF")
    for ci, esc in enumerate(["Conservador", "Base", "Optimista"], 2):
        header_cell(ws, H, ci, esc)
    header_cell(ws, H, 5, "Notas")
    ws.column_dimensions["A"].width = 42
    for col in "BCD":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 50

    ws.cell(H, 1).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws.cell(H, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(H, 1).border = BORDER
    ws.row_dimensions[H].height = 28

    PEDIDOS = ["Pedidos_mes_conservador", "Pedidos_mes_base", "Pedidos_mes_optimista"]
    ROAS = ["ROAS_conservador", "ROAS_base", "ROAS_optimista"]

    rn = H + 1

    # Volumen
    section_cell(ws, rn, 1, "VOLUMEN"); rn += 1
    label_cell(ws, rn, 2 - 1, "Pedidos / mes")
    for i in range(3):
        calc_cell(ws, rn, 2 + i, f"={PEDIDOS[i]}", '0', bold=True)
    note_cell(ws, rn, 5, "Editar en 00_SUPUESTOS · sección VOLUMEN.")
    PED_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "AOV (ticket medio con IVA)")
    for i in range(3):
        calc_cell(ws, rn, 2 + i, "=AOV", '€ #,##0.00')
    note_cell(ws, rn, 5, "Mismo AOV en los 3 escenarios.")
    rn += 1

    rn += 1
    # Ingresos
    section_cell(ws, rn, 1, "INGRESOS"); rn += 1

    label_cell(ws, rn, 1, "Ingresos brutos (con IVA)")
    for i in range(3):
        calc_cell(ws, rn, 2 + i, f"={get_column_letter(2+i)}{PED_ROW}*AOV", '€ #,##0.00', bold=True)
    note_cell(ws, rn, 5, "Pedidos × AOV.")
    ING_BRUTOS = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) IVA repercutido")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{ING_BRUTOS}*IVA/(1+IVA)", '€ #,##0.00')
    note_cell(ws, rn, 5, "21% sobre ingresos brutos. No es ingreso, es de Hacienda.")
    rn += 1

    label_cell(ws, rn, 1, "= Ingresos netos (sin IVA)")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{ING_BRUTOS}/(1+IVA)", '€ #,##0.00', bold=True)
    note_cell(ws, rn, 5, "Base imponible.")
    ING_NETOS = rn
    rn += 1

    rn += 1
    # Coste mercancía (a partir del catálogo: margen medio)
    section_cell(ws, rn, 1, "COSTE MERCANCÍA"); rn += 1

    label_cell(ws, rn, 1, "Margen bruto medio del catálogo (sobre PVP sin IVA)")
    # Calculado desde la hoja 20260508 -Todos: AVG(Margen %)
    f = f"=AVERAGE('{TODOS_SHEET}'!H3:H1000)/100"
    for i in range(3):
        calc_cell(ws, rn, 2 + i, f, '0.0%')
    note_cell(ws, rn, 5, "Promedio simple de Margen % de la hoja del catálogo.")
    MARGEN_PCT_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Coste mercancía")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{ING_NETOS}*(1-{col}{MARGEN_PCT_ROW})", '€ #,##0.00')
    note_cell(ws, rn, 5, "Ingresos netos × (1 − margen bruto medio).")
    COSTE_MERC_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "= Margen bruto")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{ING_NETOS}+{col}{COSTE_MERC_ROW}", '€ #,##0.00', bold=True)
    note_cell(ws, rn, 5, "Ingresos netos − coste mercancía.")
    MARGEN_BRUTO_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "    Margen bruto %")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{MARGEN_BRUTO_ROW}/{col}{ING_NETOS}", '0.0%')
    rn += 1

    rn += 1
    # Costes variables
    section_cell(ws, rn, 1, "COSTES VARIABLES"); rn += 1

    label_cell(ws, rn, 1, "(-) Comisiones de pago")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f"=-({col}{ING_BRUTOS}*Comision_pago_pct + {col}{PED_ROW}*Comision_pago_fija)",
                  '€ #,##0.00')
    note_cell(ws, rn, 5, "% sobre venta + fija/pedido.")
    COM_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Logística salida")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{PED_ROW}*Coste_medio_envio_pedido", '€ #,##0.00')
    note_cell(ws, rn, 5, "Pedidos × coste medio mezclado del catálogo.")
    LOG_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Devoluciones")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f"=-{col}{PED_ROW}*Tasa_devolucion*Coste_devolucion",
                  '€ #,##0.00')
    note_cell(ws, rn, 5, "Pedidos × tasa × coste por devolución.")
    DEV_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Incidencias transporte")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f"=-{col}{PED_ROW}*Tasa_incidencia*Coste_incidencia",
                  '€ #,##0.00')
    note_cell(ws, rn, 5, "Pedidos × tasa × coste por incidencia.")
    INC_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "= Margen contributivo (antes de marketing)")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f"={col}{MARGEN_BRUTO_ROW}+{col}{COM_ROW}+{col}{LOG_ROW}+{col}{DEV_ROW}+{col}{INC_ROW}",
                  '€ #,##0.00', bold=True)
    note_cell(ws, rn, 5, "Lo que queda para pagar marketing y costes fijos.")
    MC_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "    Margen contributivo %")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{MC_ROW}/{col}{ING_NETOS}", '0.0%')
    rn += 1

    rn += 1
    # Marketing
    section_cell(ws, rn, 1, "MARKETING"); rn += 1

    label_cell(ws, rn, 1, "ROAS objetivo")
    for i in range(3):
        calc_cell(ws, rn, 2 + i, f"={ROAS[i]}", '0.0"x"')
    ROAS_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Inversión marketing")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{ING_BRUTOS}/{col}{ROAS_ROW}", '€ #,##0.00')
    note_cell(ws, rn, 5, "Ingresos brutos / ROAS objetivo.")
    MK_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "    % marketing sobre ventas brutas")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{MK_ROW}/{col}{ING_BRUTOS}", '0.0%')
    rn += 1

    label_cell(ws, rn, 1, "= EBITDA tras marketing")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{MC_ROW}+{col}{MK_ROW}", '€ #,##0.00', bold=True)
    EBITDA_ROW = rn
    rn += 1

    rn += 1
    # Costes fijos
    section_cell(ws, rn, 1, "COSTES FIJOS"); rn += 1

    label_cell(ws, rn, 1, "(-) Costes fijos mensuales")
    for i in range(3):
        calc_cell(ws, rn, 2 + i, "=-TOTAL_FIJOS_MES", '€ #,##0.00')
    note_cell(ws, rn, 5, "Shopify + dominio + apps + personal (en 00_SUPUESTOS).")
    FIJOS_ROW = rn
    rn += 1

    rn += 1
    # Beneficio operativo
    label_cell(ws, rn, 1, "BENEFICIO OPERATIVO MENSUAL")
    ws.cell(rn, 1).font = Font(size=12, bold=True)
    for i in range(3):
        col = get_column_letter(2 + i)
        c = ws.cell(rn, 2 + i, value=f"={col}{EBITDA_ROW}+{col}{FIJOS_ROW}")
        c.fill = PatternFill("solid", fgColor=COLOR_KPI)
        c.font = Font(size=12, bold=True, color="C65911")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = Border(left=MED, right=MED, top=MED, bottom=MED)
        c.number_format = '€ #,##0.00;[Red]-€ #,##0.00'
    BENEF_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "    Beneficio neto %")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=IFERROR({col}{BENEF_ROW}/{col}{ING_NETOS},0)", '0.0%')
    rn += 1

    rn += 2
    # KPIs y break-even
    section_cell(ws, rn, 1, "KPIs Y BREAK-EVEN"); rn += 1

    label_cell(ws, rn, 1, "Margen contributivo €/pedido")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{MC_ROW}/{col}{PED_ROW}", '€ #,##0.00')
    MC_PED_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "Pedidos break-even mensual (cubre fijos)")
    for i in range(3):
        col = get_column_letter(2 + i)
        # fijos / (mc/pedido + marketing/pedido_negativo? no, marketing depende de ventas)
        # Mejor: pedidos para que MC * pedidos = costes fijos + marketing
        # marketing = ingresos / ROAS = (pedidos*AOV) / ROAS
        # MC neto/pedido = MC/pedido - AOV/ROAS  (lo que queda tras marketing por cada pedido)
        # break-even: pedidos = fijos / MC_neto/pedido
        f = (f"=IFERROR(TOTAL_FIJOS_MES/MAX({col}{MC_ROW}/{col}{PED_ROW}-AOV/{col}{ROAS_ROW},0.01),0)")
        kpi_cell(ws, rn, 2 + i, f, '0')
    note_cell(ws, rn, 5, "Pedidos mínimos para que ingresos > costes fijos + marketing al ROAS objetivo.")
    rn += 1

    label_cell(ws, rn, 1, "Sesiones web necesarias / mes")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{PED_ROW}/Tasa_conversion", '#,##0')
    note_cell(ws, rn, 5, "Pedidos / tasa de conversión.")
    rn += 1

    label_cell(ws, rn, 1, "CPA medio (coste por adquisición de cliente)")
    for i in range(3):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{MK_ROW}/{col}{PED_ROW}", '€ #,##0.00')
    note_cell(ws, rn, 5, "Inversión marketing / pedidos. Comparar con margen contributivo €.")
    rn += 1

    ws.freeze_panes = "B5"


# ── 02_UNIT_ECONOMICS_SKU ────────────────────────────────────────────────────

UE_HEADERS = [
    ("Proveedor", 10),
    ("SKU", 14),
    ("Producto", 38),
    ("Coste sin IVA €", 13),
    ("PVP sin IVA €", 13),
    ("PVP con IVA €", 13),
    ("Margen bruto €", 13),
    ("Margen bruto %", 13),
    ("Comisión pago €", 13),
    ("Logística €", 13),
    ("Devoluciones €", 13),
    ("Incidencias €", 13),
    ("Margen contrib. €", 14),
    ("Margen contrib. %", 14),
    ("CAC máximo €", 13),
    ("ROAS mínimo", 12),
    ("Categoría", 14),
]


def write_unit_economics(wb):
    name = "02_UNIT_ECONOMICS_SKU"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    title_cell(ws, 1, 1, "02 · UNIT ECONOMICS POR SKU — qué deja cada producto realmente",
               span=len(UE_HEADERS))
    note_cell(ws, 2, 1, "Una fila por SKU. Edita supuestos en 00_SUPUESTOS para recalcular todo.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(UE_HEADERS))

    H = 4
    for ci, (h, w) in enumerate(UE_HEADERS, 1):
        header_cell(ws, H, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[H].height = 38
    ws.freeze_panes = f"D{H+1}"

    # Leer todos los SKUs de '20260508 -Todos '
    src = wb[TODOS_SHEET]
    rn = H + 1
    for r in range(3, src.max_row + 1):
        prov = src.cell(r, 1).value
        sku = src.cell(r, 3).value
        prod = src.cell(r, 4).value
        coste = src.cell(r, 5).value
        pvp_iva = src.cell(r, 6).value      # PVP con IVA = PVP × 1.21
        pvp_rec = src.cell(r, 9).value      # PVP sin IVA
        if not prov or not sku:
            continue
        if not isinstance(coste, (int, float)) or not isinstance(pvp_rec, (int, float)):
            continue

        ws.cell(rn, 1, value=prov)
        ws.cell(rn, 2, value=sku)
        ws.cell(rn, 3, value=prod)
        ws.cell(rn, 4, value=coste).number_format = '€ #,##0.00'
        ws.cell(rn, 5, value=pvp_rec).number_format = '€ #,##0.00'
        # PVP con IVA por fórmula (consistencia)
        ws.cell(rn, 6, value=f"=E{rn}*(1+IVA)").number_format = '€ #,##0.00'
        # Margen bruto € y %
        ws.cell(rn, 7, value=f"=E{rn}-D{rn}").number_format = '€ #,##0.00'
        ws.cell(rn, 8, value=f"=IFERROR(G{rn}/E{rn},0)").number_format = '0.0%'
        # Comisión pago: % sobre PVP con IVA + fija
        ws.cell(rn, 9, value=f"=F{rn}*Comision_pago_pct+Comision_pago_fija").number_format = '€ #,##0.00'
        # Logística: para Hevea, si PVP con IVA > umbral → 0; sino → coste bajo. Balliu siempre coste.
        log_formula = (
            f'=IF(A{rn}="Hevea",'
            f'IF(F{rn}>=Hevea_umbral_envio_gratis,0,Hevea_coste_envio_bajo),'
            f'Balliu_coste_envio)'
        )
        ws.cell(rn, 10, value=log_formula).number_format = '€ #,##0.00'
        # Devoluciones (esperado por unidad)
        ws.cell(rn, 11, value=f"=Tasa_devolucion*Coste_devolucion").number_format = '€ #,##0.00'
        # Incidencias
        ws.cell(rn, 12, value=f"=Tasa_incidencia*Coste_incidencia").number_format = '€ #,##0.00'
        # Margen contributivo € = Margen bruto - comisión - logística - dev - incid
        ws.cell(rn, 13, value=f"=G{rn}-I{rn}-J{rn}-K{rn}-L{rn}").number_format = '€ #,##0.00'
        # Margen contributivo %
        ws.cell(rn, 14, value=f"=IFERROR(M{rn}/E{rn},0)").number_format = '0.0%'
        # CAC máximo = MC € (lo que puedo gastar en ads sin perder dinero, antes de fijos)
        ws.cell(rn, 15, value=f"=MAX(M{rn},0)").number_format = '€ #,##0.00'
        # ROAS mínimo = PVP con IVA / CAC máximo (a partir de ese ROAS, pierdo dinero)
        ws.cell(rn, 16, value=f"=IFERROR(F{rn}/O{rn},0)").number_format = '0.0"x"'
        # Categoría
        cat = (
            f'=IF(N{rn}>=Umbral_PUSH_pct,'
            f'IF(M{rn}>=Umbral_PUSH_eur,"🟢 PUSH","🟡 NEUTRAL"),'
            f'IF(N{rn}>=Umbral_NEUTRAL_pct,"🟡 NEUTRAL",'
            f'IF(N{rn}>=Umbral_WATCH_pct,"🟠 WATCH","🔴 NO ANUNCIAR")))'
        )
        ws.cell(rn, 17, value=cat)

        # Estilos por fila (alternar)
        fill_color = "EEF5FF" if rn % 2 == 0 else "DDEEFF"
        for ci in range(1, len(UE_HEADERS) + 1):
            c = ws.cell(rn, ci)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="right" if ci >= 4 else "left",
                                    vertical="center")
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=fill_color)
        rn += 1

    last = rn - 1

    # Filtros y formato condicional sobre Margen contrib. % (col N)
    if last >= H + 1:
        ws.auto_filter.ref = f"A{H}:{get_column_letter(len(UE_HEADERS))}{last}"
        n_col = get_column_letter(14)
        ws.conditional_formatting.add(
            f"{n_col}{H+1}:{n_col}{last}",
            ColorScaleRule(
                start_type="num", start_value="0",  start_color="F8696B",
                mid_type="num",   mid_value="0.2",  mid_color="FFEB84",
                end_type="num",   end_value="0.4",  end_color="63BE7B",
            )
        )

    # Fila TOTAL/MEDIA
    rn = last + 1
    ws.cell(rn, 3, value="TOTAL / MEDIA").font = Font(bold=True)
    ws.cell(rn, 3).fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    ws.cell(rn, 3).alignment = Alignment(horizontal="right", vertical="center")
    for ci, (kind, fmt) in [
        (4, ("AVG", '€ #,##0.00')), (5, ("AVG", '€ #,##0.00')), (6, ("AVG", '€ #,##0.00')),
        (7, ("AVG", '€ #,##0.00')), (8, ("AVG", '0.0%')),
        (13, ("AVG", '€ #,##0.00')), (14, ("AVG", '0.0%')),
        (15, ("AVG", '€ #,##0.00')), (16, ("AVG", '0.0"x"')),
    ]:
        col = get_column_letter(ci)
        f = f"=AVERAGE({col}{H+1}:{col}{last})"
        c = ws.cell(rn, ci, value=f)
        c.number_format = fmt
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER

    return last - H  # nº filas datos


# ── 03_ESCENARIOS_MARKETING ──────────────────────────────────────────────────

def write_escenarios(wb):
    name = "03_ESCENARIOS_MARKETING"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    title_cell(ws, 1, 1, "03 · ESCENARIOS DE MARKETING — cuánto invertir y qué esperar", span=6)
    note_cell(ws, 2, 1, "Modo A: dado un presupuesto, qué pedidos/tráfico/AOV necesito. Modo B: dado un ROAS, cuánto puedo invertir.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Modo A: 4 escenarios de inversión × supuestos
    H = 4
    section_cell(ws, H, 1, "MODO A — DESDE INVERSIÓN", span=6)
    H += 1

    headers_A = ["Concepto", "500 € /mes", "1.500 € /mes", "3.000 € /mes", "6.000 € /mes", "Notas"]
    for ci, h in enumerate(headers_A, 1):
        header_cell(ws, H, ci, h)
    ws.column_dimensions["A"].width = 42
    for col in "BCDE":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["F"].width = 50
    ws.row_dimensions[H].height = 28

    INV = [500, 1500, 3000, 6000]
    rn = H + 1

    # Inversión (editable)
    label_cell(ws, rn, 1, "Inversión marketing / mes")
    for i, v in enumerate(INV):
        editable_cell(ws, rn, 2 + i, v, '€ #,##0.00')
    note_cell(ws, rn, 6, "Editar para simular otros presupuestos.")
    INV_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "ROAS objetivo (escenario BASE)")
    for i in range(4):
        calc_cell(ws, rn, 2 + i, "=ROAS_base", '0.0"x"')
    note_cell(ws, rn, 6, "Por defecto ROAS_base. Editable en 00_SUPUESTOS.")
    ROAS_A_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "Ingresos brutos generados (con IVA)")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f"={col}{INV_ROW}*{col}{ROAS_A_ROW}", '€ #,##0.00', bold=True)
    note_cell(ws, rn, 6, "Inversión × ROAS.")
    ING_A_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "Pedidos generados")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{ING_A_ROW}/AOV", '0')
    note_cell(ws, rn, 6, "Ingresos / AOV.")
    PED_A_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "Sesiones web necesarias")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{PED_A_ROW}/Tasa_conversion", '#,##0')
    note_cell(ws, rn, 6, "Pedidos / tasa de conversión.")
    rn += 1

    label_cell(ws, rn, 1, "CPA medio")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"={col}{INV_ROW}/{col}{PED_A_ROW}", '€ #,##0.00')
    note_cell(ws, rn, 6, "Inversión / pedidos.")
    rn += 1

    label_cell(ws, rn, 1, "Margen contributivo total (antes de marketing)")
    for i in range(4):
        col = get_column_letter(2 + i)
        # MC unitario: AOV - coste mercancía (ya con IVA repartido) - var/pedido
        # Aprox: (PVP_neto × margen_pct_medio) - VAR_COST_PER_ORDER
        f = (f"={col}{PED_A_ROW}*"
             f"((AOV/(1+IVA))*AVERAGE('{TODOS_SHEET}'!H3:H1000)/100"
             f"-VAR_COST_PER_ORDER)")
        calc_cell(ws, rn, 2 + i, f, '€ #,##0.00')
    note_cell(ws, rn, 6, "Pedidos × (margen bruto unitario − coste variable unitario).")
    MC_A_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "(-) Inversión marketing")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=-{col}{INV_ROW}", '€ #,##0.00')
    rn += 1

    label_cell(ws, rn, 1, "(-) Costes fijos mensuales")
    for i in range(4):
        calc_cell(ws, rn, 2 + i, "=-TOTAL_FIJOS_MES", '€ #,##0.00')
    rn += 1

    label_cell(ws, rn, 1, "= Beneficio operativo")
    ws.cell(rn, 1).font = Font(bold=True, size=11)
    for i in range(4):
        col = get_column_letter(2 + i)
        c = ws.cell(rn, 2 + i, value=f"={col}{MC_A_ROW}-{col}{INV_ROW}-TOTAL_FIJOS_MES")
        c.fill = PatternFill("solid", fgColor=COLOR_KPI)
        c.font = Font(size=11, bold=True, color="C65911")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = Border(left=MED, right=MED, top=MED, bottom=MED)
        c.number_format = '€ #,##0.00;[Red]-€ #,##0.00'
    rn += 1

    rn += 2
    section_cell(ws, rn, 1, "MODO B — DESDE ROAS OBJETIVO", span=6)
    rn += 1

    headers_B = ["Concepto", "ROAS 3x", "ROAS 4x", "ROAS 5x", "ROAS 6x", "Notas"]
    for ci, h in enumerate(headers_B, 1):
        header_cell(ws, rn, ci, h)
    ws.row_dimensions[rn].height = 28
    rn += 1

    label_cell(ws, rn, 1, "ROAS objetivo")
    for i, v in enumerate([3, 4, 5, 6]):
        editable_cell(ws, rn, 2 + i, v, '0.0"x"')
    note_cell(ws, rn, 6, "Editar para evaluar otros ROAS.")
    ROAS_B_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "Margen contributivo €/pedido (medio catálogo)")
    f = (f"=(AOV/(1+IVA))*AVERAGE('{TODOS_SHEET}'!H3:H1000)/100"
         f"-VAR_COST_PER_ORDER")
    for i in range(4):
        calc_cell(ws, rn, 2 + i, f, '€ #,##0.00')
    note_cell(ws, rn, 6, "Margen bruto unitario − coste variable unitario.")
    MC_PED_B = rn
    rn += 1

    label_cell(ws, rn, 1, "CAC máximo aceptable / pedido")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i, f"=AOV/{col}{ROAS_B_ROW}", '€ #,##0.00')
    note_cell(ws, rn, 6, "AOV / ROAS objetivo.")
    CAC_B_ROW = rn
    rn += 1

    label_cell(ws, rn, 1, "¿Es rentable a este ROAS?")
    for i in range(4):
        col = get_column_letter(2 + i)
        calc_cell(ws, rn, 2 + i,
                  f'=IF({col}{MC_PED_B}>{col}{CAC_B_ROW},"✅ SÍ","❌ NO")', None)
        ws.cell(rn, 2 + i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(rn, 2 + i).font = Font(size=11, bold=True)
    note_cell(ws, rn, 6, "MC/pedido > CAC máximo → invertir.")
    rn += 1

    label_cell(ws, rn, 1, "Inversión sostenible / mes (con beneficio ≥ 0)")
    for i in range(4):
        col = get_column_letter(2 + i)
        # Beneficio ≥ 0 → pedidos × MC_pedido − pedidos × CAC − fijos ≥ 0
        # → pedidos × (MC − CAC) ≥ fijos
        # → pedidos_min = fijos / (MC − CAC)
        # → inversión = pedidos × CAC
        # Pero si MC < CAC, no hay break-even.
        f = (f"=IF({col}{MC_PED_B}>{col}{CAC_B_ROW},"
             f"TOTAL_FIJOS_MES/({col}{MC_PED_B}-{col}{CAC_B_ROW})*{col}{CAC_B_ROW},0)")
        calc_cell(ws, rn, 2 + i, f, '€ #,##0.00')
    note_cell(ws, rn, 6, "Inversión a partir de la cual cubrimos costes fijos. 0 = no rentable.")
    rn += 1

    label_cell(ws, rn, 1, "Pedidos break-even")
    for i in range(4):
        col = get_column_letter(2 + i)
        f = (f"=IF({col}{MC_PED_B}>{col}{CAC_B_ROW},"
             f"TOTAL_FIJOS_MES/({col}{MC_PED_B}-{col}{CAC_B_ROW}),0)")
        calc_cell(ws, rn, 2 + i, f, '0')
    rn += 1

    label_cell(ws, rn, 1, "Ingresos brutos break-even")
    for i in range(4):
        col = get_column_letter(2 + i)
        f = (f"=IF({col}{MC_PED_B}>{col}{CAC_B_ROW},"
             f"TOTAL_FIJOS_MES/({col}{MC_PED_B}-{col}{CAC_B_ROW})*AOV,0)")
        calc_cell(ws, rn, 2 + i, f, '€ #,##0.00')
    rn += 1

    ws.freeze_panes = "B5"


# ── 04_PRODUCTOS_PRIORIDAD ───────────────────────────────────────────────────

def write_prioridad(wb):
    name = "04_PRODUCTOS_PRIORIDAD"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    title_cell(ws, 1, 1, "04 · PRODUCTOS PRIORIDAD — top 30 por margen contributivo €", span=8)
    note_cell(ws, 2, 1, "Vista filtrable de la hoja 02_UNIT_ECONOMICS_SKU. Útil para decidir qué empujar en ads y SEO.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["#", "Proveedor", "SKU", "Producto",
               "PVP con IVA €", "Margen contrib. €", "Margen contrib. %", "Categoría"]
    widths = [5, 10, 14, 38, 14, 16, 16, 14]
    H = 4
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws, H, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[H].height = 28
    ws.freeze_panes = f"A{H+1}"

    # Top 30 por margen contributivo € — fórmula con LARGE
    # Referencias a 02_UNIT_ECONOMICS_SKU (col M = MC €, col N = MC %)
    UE = "02_UNIT_ECONOMICS_SKU"
    # rangos: A=Prov, B=SKU, C=Prod, F=PVP_iva, M=MC€, N=MC%, Q=Categoría
    for i in range(30):
        rn = H + 1 + i
        rank = i + 1
        ws.cell(rn, 1, value=rank)
        ws.cell(rn, 1).font = Font(bold=True, size=10)
        ws.cell(rn, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(rn, 1).border = BORDER

        # Buscar el rank-ésimo mayor MC € y traer el resto via INDEX/MATCH
        # Pero MATCH puede fallar si hay duplicados. Mejor usar LARGE + INDEX agrupado.
        # Para evitar duplicados con LARGE: usar (MC + ROW/1e6) como tiebreaker.
        # Más robusto: SMALL/LARGE con un array uno-a-uno. En Excel viejo es complejo.
        # Aproximación: usar LARGE y MATCH; en caso de empate, INDEX devuelve la 1ª.
        # Para este reporte es suficiente.
        rank_formula = f"=LARGE({UE}!$M$5:$M$1000,{rank})"
        ws.cell(rn, 6, value=rank_formula).number_format = '€ #,##0.00'

        # MATCH para encontrar la fila del MC
        match = f"MATCH(LARGE({UE}!$M$5:$M$1000,{rank}),{UE}!$M$5:$M$1000,0)+4"
        ws.cell(rn, 2, value=f"=INDEX({UE}!$A:$A,{match})")
        ws.cell(rn, 3, value=f"=INDEX({UE}!$B:$B,{match})")
        ws.cell(rn, 4, value=f"=INDEX({UE}!$C:$C,{match})")
        ws.cell(rn, 5, value=f"=INDEX({UE}!$F:$F,{match})").number_format = '€ #,##0.00'
        ws.cell(rn, 7, value=f"=INDEX({UE}!$N:$N,{match})").number_format = '0.0%'
        ws.cell(rn, 8, value=f"=INDEX({UE}!$Q:$Q,{match})")
        ws.cell(rn, 8).alignment = Alignment(horizontal="center", vertical="center")

        # Estilos
        fill_color = "EEF5FF" if rn % 2 == 0 else "DDEEFF"
        for ci in range(1, 9):
            c = ws.cell(rn, ci)
            if ci != 1:  # col 1 (rank) tiene su propio estilo bold
                c.font = Font(size=9)
                c.alignment = Alignment(
                    horizontal="right" if ci in (5, 6, 7) else "left",
                    vertical="center")
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=fill_color)


# ── 05_DASHBOARD ─────────────────────────────────────────────────────────────

def write_dashboard(wb):
    name = "05_DASHBOARD"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    title_cell(ws, 1, 1, "05 · DASHBOARD SANTAVILA — visión 1 página", span=6, size=16)
    note_cell(ws, 2, 1, "KPIs del modelo financiero. Editar valores en 00_SUPUESTOS para recalcular.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 18

    UE = "02_UNIT_ECONOMICS_SKU"
    PNL = "01_PNL_SANTAVILA"
    rn = 4

    # CATÁLOGO
    section_cell(ws, rn, 2, "🛒 CATÁLOGO", span=2)
    section_cell(ws, rn, 5, "💰 P&L MENSUAL — escenario BASE", span=2)
    rn += 1

    # SKUs total
    label_cell(ws, rn, 2, "Total SKUs en catálogo", indent=2)
    kpi_cell(ws, rn, 3, f"=COUNTA({UE}!$B$5:$B$1000)", '#,##0')

    # Ingresos brutos base
    label_cell(ws, rn, 5, "Ingresos brutos / mes (con IVA)", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"Ingresos brutos (con IVA)\",{PNL}!$A:$A,0))", '€ #,##0.00')
    rn += 1

    label_cell(ws, rn, 2, "Margen contrib. medio €/unidad", indent=2)
    kpi_cell(ws, rn, 3, f"=AVERAGE({UE}!$M$5:$M$1000)", '€ #,##0.00')

    label_cell(ws, rn, 5, "Margen contributivo / mes", indent=2)
    # MC base: usar fórmula del P&L
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"= Margen contributivo (antes de marketing)\",{PNL}!$A:$A,0))", '€ #,##0.00')
    rn += 1

    label_cell(ws, rn, 2, "Margen contrib. medio %", indent=2)
    kpi_cell(ws, rn, 3, f"=AVERAGE({UE}!$N$5:$N$1000)", '0.0%')

    label_cell(ws, rn, 5, "Inversión marketing / mes", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"(-) Inversión marketing\",{PNL}!$A:$A,0))", '€ #,##0.00')
    rn += 1

    label_cell(ws, rn, 2, "% SKUs PUSH (margen alto)", indent=2)
    kpi_cell(ws, rn, 3, f'=COUNTIF({UE}!$Q$5:$Q$1000,"*PUSH*")/COUNTA({UE}!$B$5:$B$1000)', '0.0%')

    label_cell(ws, rn, 5, "Beneficio operativo / mes", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"BENEFICIO OPERATIVO MENSUAL\",{PNL}!$A:$A,0))", '€ #,##0.00;[Red]-€ #,##0.00')
    rn += 2

    # Distribución por categoría
    section_cell(ws, rn, 2, "📊 DISTRIBUCIÓN POR CATEGORÍA", span=2)
    section_cell(ws, rn, 5, "🎯 BREAK-EVEN", span=2)
    rn += 1

    for cat, color in [("🟢 PUSH", "C6EFCE"), ("🟡 NEUTRAL", "FFEB9C"),
                       ("🟠 WATCH", "F8CBAD"), ("🔴 NO ANUNCIAR", "FFC7CE")]:
        label_cell(ws, rn, 2, cat, indent=2)
        c = ws.cell(rn, 3, value=f'=COUNTIF({UE}!$Q$5:$Q$1000,"*{cat[2:]}*")')
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER
        c.number_format = '#,##0'
        rn += 1

    rn -= 4  # rebobinar para escribir break-even en columnas E-F al lado
    label_cell(ws, rn, 5, "Pedidos break-even (escenario base)", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"Pedidos break-even mensual (cubre fijos)\",{PNL}!$A:$A,0))", '0')
    rn += 1
    label_cell(ws, rn, 5, "Sesiones web necesarias / mes", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"Sesiones web necesarias / mes\",{PNL}!$A:$A,0))", '#,##0')
    rn += 1
    label_cell(ws, rn, 5, "CPA medio (escenario base)", indent=2)
    kpi_cell(ws, rn, 6, f"=INDEX({PNL}!$C:$C,MATCH(\"CPA medio (coste por adquisición de cliente)\",{PNL}!$A:$A,0))", '€ #,##0.00')
    rn += 1
    label_cell(ws, rn, 5, "ROAS objetivo escenario BASE", indent=2)
    kpi_cell(ws, rn, 6, "=ROAS_base", '0.0"x"')
    rn += 1

    rn += 2
    section_cell(ws, rn, 2, "🏆 TOP 5 PRODUCTOS POR MARGEN CONTRIBUTIVO €", span=5)
    rn += 1
    for i, (h, w) in enumerate([("#", 5), ("SKU", 14), ("Producto", 38),
                                ("PVP €", 12), ("MC €", 14), ("MC %", 12)]):
        header_cell(ws, rn, 2 + i if i > 0 else 2, h)
    rn += 1
    for i in range(5):
        rank = i + 1
        ws.cell(rn, 2, value=rank).font = Font(bold=True, size=10)
        ws.cell(rn, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(rn, 2).border = BORDER
        match = f"MATCH(LARGE({UE}!$M$5:$M$1000,{rank}),{UE}!$M$5:$M$1000,0)+4"
        ws.cell(rn, 3, value=f"=INDEX({UE}!$B:$B,{match})")
        ws.cell(rn, 4, value=f"=INDEX({UE}!$C:$C,{match})")
        ws.cell(rn, 5, value=f"=INDEX({UE}!$F:$F,{match})").number_format = '€ #,##0.00'
        ws.cell(rn, 6, value=f"=LARGE({UE}!$M$5:$M$1000,{rank})").number_format = '€ #,##0.00'
        ws.cell(rn, 7, value=f"=INDEX({UE}!$N:$N,{match})").number_format = '0.0%'
        for ci in range(2, 8):
            c = ws.cell(rn, ci)
            c.border = BORDER
            if ci != 2:  # col 2 (rank) ya tiene font bold
                c.font = Font(size=10)
        rn += 1


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not XLSX.exists():
        sys.exit(f"No existe {XLSX}")

    BACKUPS.mkdir(exist_ok=True)
    bk = BACKUPS / f"Santavila_pre_pnl_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(XLSX, bk)
    print(f"✓ Backup: {bk.relative_to(BASE)}")

    print(f"\nAbriendo {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX)
    print(f"  Hojas existentes: {wb.sheetnames}")
    if TODOS_SHEET not in wb.sheetnames:
        sys.exit(f"  ✗ No existe '{TODOS_SHEET}'")

    print("\n→ Escribiendo 00_SUPUESTOS...")
    write_supuestos(wb)
    print("→ Escribiendo 01_PNL_SANTAVILA...")
    write_pnl(wb)
    print("→ Escribiendo 02_UNIT_ECONOMICS_SKU...")
    n_ue = write_unit_economics(wb)
    print(f"   {n_ue} SKUs procesados")
    print("→ Escribiendo 03_ESCENARIOS_MARKETING...")
    write_escenarios(wb)
    print("→ Escribiendo 04_PRODUCTOS_PRIORIDAD...")
    write_prioridad(wb)
    print("→ Escribiendo 05_DASHBOARD...")
    write_dashboard(wb)

    wb.save(XLSX)
    print(f"\n✅ {XLSX.name} actualizado")
    print(f"   Hojas finales: {wb.sheetnames}")


if __name__ == "__main__":
    main()
