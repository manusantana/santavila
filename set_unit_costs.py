#!/usr/bin/env python3
"""
set_unit_costs.py

Rellena el campo "Coste por artículo" (unitCost) en Shopify para los productos
donde el auditor usaba estimaciones del Excel porque Shopify no tenía el dato.

Estrategia de emparejamiento:
  Excel fila → precio psicológico (col G) → busca variante Shopify con ese precio exacto
  → aplica el coste (col E) de esa fila.

Handles procesados:
  - balliu-mesa-exterior-aluminio-8080-cm-ef580ae2  (BRUNEI – 4 tamaños × 2 tableros)
  - balliu-mesa-exterior-aluminio-7070-cm-724b0db0  (CAPRI  – 3 tamaños × 2 tableros)

Modos:
  --dry-run (default)  → muestra emparejamientos sin tocar Shopify
  --apply              → aplica los costes
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

BASE   = Path(__file__).resolve().parent
XLSX   = BASE / "Santavila.xlsx"
SHEET  = "20260508 -Todos "
SHOP   = "mueblesexterior.myshopify.com"
API    = f"https://{SHOP}/admin/api/2026-01/graphql.json"

HANDLES = [
    "balliu-mesa-exterior-aluminio-8080-cm-ef580ae2",
    "balliu-mesa-exterior-aluminio-7070-cm-724b0db0",
]


def read_token() -> str:
    for fname in (".env.local", ".envlocal", ".env"):
        p = BASE / fname
        if not p.exists():
            continue
        m = re.search(r"^SHOPIFY_ACCESS_TOKEN=(.*)$", p.read_text(encoding="utf-8"), re.M)
        if m:
            return m.group(1).strip().strip('"').strip("'")
    sys.exit("✗ SHOPIFY_ACCESS_TOKEN no encontrado")


def gql(token: str, query: str, variables: dict | None = None) -> dict:
    payload = json.dumps({"query": query, "variables": variables or {}}).encode()
    for attempt in range(1, 5):
        try:
            req = urllib.request.Request(
                API, data=payload,
                headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as r:
                data = json.loads(r.read())
            if data.get("errors"):
                raise RuntimeError(json.dumps(data["errors"]))
            return data["data"]
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(float(e.headers.get("Retry-After", 2)))
                continue
            raise
        except Exception:
            time.sleep(1.5 * attempt)
    raise RuntimeError("GraphQL falló tras reintentos")


QUERY_PRODUCT = """
query($handle: String!) {
  productByHandle(handle: $handle) {
    id title status
    variants(first: 100) {
      nodes {
        id sku price title
        inventoryItem { id unitCost { amount } }
      }
    }
  }
}
"""

MUTATION_COST = """
mutation($pid: ID!, $variants: [ProductVariantsBulkInput!]!) {
  productVariantsBulkUpdate(productId: $pid, variants: $variants) {
    productVariants {
      id sku
      inventoryItem { unitCost { amount } }
    }
    userErrors { field message }
  }
}
"""


def load_excel_costs(handles: list[str]) -> dict[str, dict[float, float]]:
    """Devuelve {handle: {psy_price: coste}} leyendo col E (coste) y G (psy)."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    result: dict[str, dict[float, float]] = {}
    for r in range(3, ws.max_row + 1):
        h = str(ws.cell(r, 2).value or "").strip()
        if h not in handles:
            continue
        coste = ws.cell(r, 5).value
        psy   = ws.cell(r, 7).value
        if not isinstance(coste, (int, float)) or not isinstance(psy, (int, float)):
            continue
        coste_f = round(float(coste), 2)
        psy_f   = round(float(psy),   2)
        if h not in result:
            result[h] = {}
        if psy_f in result[h]:
            # Dos filas con el mismo PSY → ambigüedad, excluir
            result[h][psy_f] = None
        else:
            result[h][psy_f] = coste_f
    # Limpiar ambiguos
    for h in result:
        result[h] = {k: v for k, v in result[h].items() if v is not None}
    return result


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    dry = not args.apply

    token = read_token()

    cost_map = load_excel_costs(HANDLES)
    print("Costes cargados del Excel:")
    for h, m in cost_map.items():
        print(f"  {h}: {len(m)} entradas (precio→coste)")

    total_ok = total_skip = total_err = 0

    for handle in HANDLES:
        psy_to_cost = cost_map.get(handle, {})
        if not psy_to_cost:
            print(f"\n[!] Sin datos Excel para {handle}")
            continue

        data = gql(token, QUERY_PRODUCT, {"handle": handle})
        prod = data.get("productByHandle")
        if not prod:
            print(f"\n[!] Producto no encontrado en Shopify: {handle}")
            continue

        print(f"\n── {prod['title']} [{prod['status']}] ──")
        variants = prod["variants"]["nodes"]

        to_update = []
        for v in variants:
            price        = round(float(v["price"]), 2)
            current_cost = (v["inventoryItem"].get("unitCost") or {}).get("amount")
            expected_cost = psy_to_cost.get(price)

            if expected_cost is None:
                print(f"  ? {v['sku']:<45} {price:>8.2f}€  — sin match en Excel")
                total_skip += 1
                continue

            if current_cost and abs(float(current_cost) - expected_cost) < 0.01:
                print(f"  · {v['sku']:<45} {price:>8.2f}€  coste={expected_cost:.2f}€ (ya OK)")
                total_skip += 1
                continue

            mark = "[DRY]" if dry else "     "
            print(f"  {mark} {v['sku']:<45} {price:>8.2f}€  coste: {current_cost or 'n/a'} → {expected_cost:.2f}€")
            if not dry:
                to_update.append({"id": v["id"], "inventoryItem": {"cost": str(expected_cost)}})
            total_ok += 1

        if to_update and not dry:
            try:
                res = gql(token, MUTATION_COST, {"pid": prod["id"], "variants": to_update})
                errs = res["productVariantsBulkUpdate"]["userErrors"]
                if errs:
                    print(f"  ✗ userErrors: {errs}")
                    total_err += len(to_update)
                    total_ok  -= len(to_update)
                else:
                    updated = res["productVariantsBulkUpdate"]["productVariants"]
                    print(f"  ✓ {len(updated)} variantes actualizadas")
            except Exception as e:
                print(f"  ✗ Excepción: {e}")
                total_err += len(to_update)
                total_ok  -= len(to_update)

    print(f"\n── Resumen ──")
    print(f"  {'[DRY] ' if dry else ''}Actualizados: {total_ok}")
    print(f"  Sin match / ya OK: {total_skip}")
    if total_err:
        print(f"  Errores: {total_err}")
    if dry:
        print("\nPara aplicar: python3 set_unit_costs.py --apply")


if __name__ == "__main__":
    main()
