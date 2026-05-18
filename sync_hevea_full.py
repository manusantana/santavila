#!/usr/bin/env python3
"""
sync_hevea_full.py

Sincroniza TODOS los productos Hevea usando el CSV como fuente de verdad.

Fuente: proveedores_raw/hevea/20260507 ▶️CSV hevea 07_05_25.csv
Destinos:
  1. Excel Santavila.xlsx hoja "20260508 -Todos "
     - col E (coste)      = Precio exworks (sin iva)
     - col F (PVP con IVA) = PVP Recomendado (sin iva) × 1.21
     - col G (PSY)         = psy_price(col F)
     - col K (carrier)     = 50 si PVP_IVA < 500, si no 0
  2. Shopify (por handle desde col B del Excel):
     - price         = PSY(PVP × 1.21)
     - compareAtPrice = PVP × 1.21 (si es distinto de PSY)
     - unitCost      = exworks
     - body_html     = descripción del CSV + tabla de dimensiones
     - tags          = mantiene existentes

SKUs con duplicado en el CSV (no se pueden sincronizar automáticamente):
  Se reportan al final para revisión manual.

Modos:
  --dry-run (default)  → muestra cambios sin tocar nada
  --apply              → aplica Excel + Shopify
  --only-excel         → solo corrige el Excel, no Shopify
  --only-shopify       → solo actualiza Shopify
  --limit N            → procesa solo N productos (test)
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

BASE   = Path(__file__).resolve().parent
XLSX   = BASE / "Santavila.xlsx"
SHEET  = "20260508 -Todos "
CSV_FILE = BASE / "proveedores_raw/hevea/20260507 ▶️CSV hevea 07_05_25.csv"
SHOP   = "mueblesexterior.myshopify.com"
API    = f"https://{SHOP}/admin/api/2026-01/graphql.json"

COL_PROV    = 1
COL_HANDLE  = 2
COL_SKU     = 3
COL_PROD    = 4
COL_COSTE   = 5
COL_PVP     = 6
COL_PSY     = 7
COL_CARRIER = 11

IVA = 1.21
UMBRALES = (100, 150, 200, 250, 300, 350, 400, 450, 500, 600, 700, 800, 900,
            1000, 1200, 1500, 1800, 2000)


# ── Psicológico ──────────────────────────────────────────────────────────────

def _next_high_ticket(p: float) -> float:
    n = math.ceil(p)
    while n % 10 not in (0, 5, 9):
        n += 1
    return float(n)


def _round_up_to_95(p: float) -> float:
    base = math.floor(p)
    cand = base + 0.95
    if cand + 1e-9 < p:
        cand += 1
    return round(cand, 2)


def psy_price(price_bruto: float) -> float:
    if price_bruto < 50:
        return _round_up_to_95(price_bruto)
    if price_bruto <= 500:
        for u in UMBRALES:
            if u <= price_bruto <= u * 1.05:
                return round(u - 0.10, 2)
        return _round_up_to_95(price_bruto)
    return _next_high_ticket(price_bruto)


# ── Token ─────────────────────────────────────────────────────────────────────

def read_token() -> str:
    for fname in (".env.local", ".envlocal", ".env"):
        p = BASE / fname
        if not p.exists():
            continue
        m = re.search(r"^SHOPIFY_ACCESS_TOKEN=(.*)$", p.read_text(encoding="utf-8"), re.M)
        if m:
            return m.group(1).strip().strip('"').strip("'")
    sys.exit("✗ SHOPIFY_ACCESS_TOKEN no encontrado")


# ── GraphQL ───────────────────────────────────────────────────────────────────

_throttle: dict[str, float] = {"available": 2000.0}
PAUSE_THRESHOLD = 200
PAUSE_SECONDS   = 2.0


def gql(token: str, query: str, variables: dict | None = None) -> dict:
    payload = json.dumps({"query": query, "variables": variables or {}}).encode()
    last_err = None
    for attempt in range(1, 5):
        try:
            req = urllib.request.Request(
                API, data=payload,
                headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as r:
                data = json.loads(r.read())
            if data.get("errors"):
                raise RuntimeError(json.dumps(data["errors"]))
            ts = data.get("extensions", {}).get("cost", {}).get("throttleStatus", {})
            if ts:
                _throttle["available"] = ts.get("currentlyAvailable", 0)
            if _throttle["available"] < PAUSE_THRESHOLD:
                time.sleep(PAUSE_SECONDS)
            return data["data"]
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(float(e.headers.get("Retry-After", 2)))
                continue
            body = e.read().decode("utf-8", "replace")[:300]
            last_err = f"HTTP {e.code}: {body}"
        except Exception as e:
            last_err = str(e)
        time.sleep(1.5 * attempt)
    raise RuntimeError(f"GraphQL falló: {last_err}")


QUERY_PRODUCT = """
query($handle: String!) {
  productByHandle(handle: $handle) {
    id title status descriptionHtml
    variants(first: 1) {
      nodes {
        id sku price compareAtPrice
        inventoryItem { unitCost { amount } }
      }
    }
  }
}
"""

MUTATION_UPDATE = """
mutation($pid: ID!, $input: ProductInput!, $variants: [ProductVariantsBulkInput!]!) {
  productUpdate(input: $input) {
    product { id descriptionHtml }
    userErrors { field message }
  }
  productVariantsBulkUpdate(productId: $pid, variants: $variants) {
    productVariants { id price compareAtPrice inventoryItem { unitCost { amount } } }
    userErrors { field message }
  }
}
"""


# ── Carga del CSV ─────────────────────────────────────────────────────────────

def load_csv() -> tuple[dict[str, dict], set[str]]:
    """Devuelve (csv_by_sku, duplicate_skus).
    csv_by_sku: {sku: row_dict} — solo SKUs únicos.
    duplicate_skus: SKUs que aparecen más de una vez en el CSV."""
    from collections import Counter
    rows = list(csv.DictReader(open(CSV_FILE, encoding="utf-8-sig")))
    counts = Counter(r["SKU"].strip() for r in rows if r["SKU"].strip())
    dups = {sku for sku, c in counts.items() if c > 1}
    unique: dict[str, dict] = {}
    for r in rows:
        sku = r["SKU"].strip()
        if sku and sku not in dups:
            unique[sku] = r
    return unique, dups


# ── Descripción HTML ──────────────────────────────────────────────────────────

def build_html(descripcion: str, ancho: str, fondo: str, alto: str) -> str:
    parts = []
    if descripcion.strip():
        parts.append(f"<p>{descripcion.strip()}</p>")
    dims = [(d, v) for d, v in [("Ancho", ancho), ("Fondo", fondo), ("Alto", alto)]
            if str(v).strip() and str(v).strip() not in ("", "0", "xx")]
    if dims:
        rows_html = "".join(
            f"<tr><td><strong>{d}</strong></td><td>{v} cm</td></tr>" for d, v in dims
        )
        parts.append(
            f"<table><thead><tr><th>Dimensión</th><th>Medida</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table>"
        )
    return "\n".join(parts)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply",        action="store_true", help="Aplica todos los cambios")
    parser.add_argument("--only-excel",   action="store_true", help="Solo corrige el Excel")
    parser.add_argument("--only-shopify", action="store_true", help="Solo actualiza Shopify")
    parser.add_argument("--limit",        type=int, default=None)
    args = parser.parse_args()

    dry = not args.apply
    do_excel   = args.apply and not args.only_shopify
    do_shopify = args.apply and not args.only_excel
    if args.only_excel:
        do_excel, do_shopify = args.apply, False
    if args.only_shopify:
        do_excel, do_shopify = False, args.apply

    # CSV
    csv_data, dup_skus = load_csv()
    print(f"CSV cargado: {len(csv_data)} SKUs únicos, {len(dup_skus)} SKUs duplicados")
    print(f"SKUs duplicados (requieren revisión manual): {sorted(dup_skus)}")

    # Excel
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    token = read_token() if (not dry and do_shopify) or (dry) else None
    if dry:
        token = read_token()

    # Construir lista de filas Hevea
    hevea_skus = set(csv_data.keys()) | dup_skus
    hevea_rows: list[tuple[int, str, str, str]] = []  # (row, handle, sku, producto)
    seen_handles: set[str] = set()

    for r in range(3, ws.max_row + 1):
        sku    = str(ws.cell(r, COL_SKU).value    or "").strip()
        handle = str(ws.cell(r, COL_HANDLE).value or "").strip()
        prod   = str(ws.cell(r, COL_PROD).value   or "").strip()
        prov   = str(ws.cell(r, COL_PROV).value   or "").lower()

        is_hevea = ("hevea" in prov or
                    sku.startswith("557-") or sku.startswith("526-") or
                    sku.startswith("600-") or sku.startswith("123-") or
                    sku.startswith("921-") or sku.startswith("928-"))
        if not is_hevea or not handle or not sku:
            continue
        hevea_rows.append((r, handle, sku, prod))

    if args.limit:
        hevea_rows = hevea_rows[:args.limit]

    print(f"\nFilas Hevea en Excel: {len(hevea_rows)}")

    # ── Counters ──
    excel_changed = excel_ok = shopify_updated = shopify_ok = shopify_skip = 0
    unresolved: list[str] = []

    for row_idx, (r, handle, sku, prod_excel) in enumerate(hevea_rows, 1):
        # SKU duplicado → saltar + registrar
        if sku in dup_skus:
            unresolved.append(f"  {handle} | SKU={sku} (duplicado en CSV — revisión manual)")
            continue

        csv_row = csv_data.get(sku)
        if not csv_row:
            unresolved.append(f"  {handle} | SKU={sku} (no encontrado en CSV)")
            continue

        # Valores del CSV
        exworks_raw = csv_row.get("Precio exworks (sin iva)", "").strip()
        pvp_raw     = csv_row.get("PVP Recomendado (sin iva)", "").strip()
        descripcion = csv_row.get("Descripción", "").strip()
        ancho       = csv_row.get("Ancho (cm)", "").strip()
        fondo       = csv_row.get("Fondo (cm)", "").strip()
        alto        = csv_row.get("Alto (cm)", "").strip()

        try:
            exworks = float(exworks_raw)
            pvp_sin_iva = float(pvp_raw)
        except (ValueError, TypeError):
            unresolved.append(f"  {handle} | SKU={sku} (precio inválido en CSV)")
            continue

        pvp_iva = round(pvp_sin_iva * IVA, 2)
        psy     = psy_price(pvp_iva)
        carrier = 50 if pvp_iva < 500 else 0

        # ── Excel update ──
        cur_coste   = ws.cell(r, COL_COSTE).value
        cur_pvp     = ws.cell(r, COL_PVP).value
        cur_psy     = ws.cell(r, COL_PSY).value
        cur_carrier = ws.cell(r, COL_CARRIER).value

        excel_diff = []
        if not isinstance(cur_coste, (int, float)) or abs(float(cur_coste) - exworks) > 0.01:
            excel_diff.append(f"coste {cur_coste}→{exworks}")
        if not isinstance(cur_pvp, (int, float)) or abs(float(cur_pvp) - pvp_iva) > 0.01:
            excel_diff.append(f"pvp {cur_pvp}→{pvp_iva}")
        if not isinstance(cur_psy, (int, float)) or abs(float(cur_psy) - psy) > 0.01:
            excel_diff.append(f"psy {cur_psy}→{psy}")
        if cur_carrier != carrier:
            excel_diff.append(f"carrier {cur_carrier}→{carrier}")

        if excel_diff:
            mark = "     " if do_excel else "[DRY]"
            print(f"  {mark} EXCEL r{r:3d} {handle[:50]:<50} | {' · '.join(excel_diff)}")
            if do_excel:
                ws.cell(r, COL_COSTE).value   = exworks
                ws.cell(r, COL_PVP).value     = pvp_iva
                ws.cell(r, COL_PSY).value     = psy
                ws.cell(r, COL_CARRIER).value = carrier
            excel_changed += 1
        else:
            excel_ok += 1

        # ── Shopify update ──
        if not token:
            continue

        data = gql(token, QUERY_PRODUCT, {"handle": handle})
        prod = data.get("productByHandle")
        if not prod:
            unresolved.append(f"  {handle} | SKU={sku} (no existe en Shopify)")
            continue
        if prod["status"] == "DRAFT":
            shopify_skip += 1
            continue

        v = prod["variants"]["nodes"][0]
        cur_price   = float(v["price"])
        cur_compare = float(v["compareAtPrice"] or 0)
        cur_cost    = float((v["inventoryItem"].get("unitCost") or {}).get("amount") or 0)

        new_html = build_html(descripcion, ancho, fondo, alto)

        shopify_diff = []
        if abs(cur_price - psy) > 0.01:
            shopify_diff.append(f"price {cur_price}→{psy}")
        # compareAtPrice solo cuando pvp_iva > psy (precio rebajado, tachado tiene sentido)
        target_compare = pvp_iva if pvp_iva > psy + 0.01 else 0
        if target_compare and abs(cur_compare - target_compare) > 0.01:
            shopify_diff.append(f"compare {cur_compare}→{target_compare}")
        elif not target_compare and cur_compare > 0:
            shopify_diff.append(f"compare {cur_compare}→(eliminar)")
        if abs(cur_cost - exworks) > 0.01:
            shopify_diff.append(f"cost {cur_cost}→{exworks}")
        if new_html and (prod["descriptionHtml"] or "").strip() != new_html.strip():
            shopify_diff.append("descripción")

        if shopify_diff:
            mark = "     " if do_shopify else "[DRY]"
            print(f"  {mark} SHOP  {handle[:50]:<50} | {' · '.join(shopify_diff)}")
            if do_shopify:
                variant_input: dict = {"id": v["id"]}
                if abs(cur_price - psy) > 0.01:
                    variant_input["price"] = str(psy)
                if target_compare and abs(cur_compare - target_compare) > 0.01:
                    variant_input["compareAtPrice"] = str(target_compare)
                elif not target_compare and cur_compare > 0:
                    variant_input["compareAtPrice"] = None
                variant_input["inventoryItem"] = {"cost": str(exworks)}

                product_input: dict = {"id": prod["id"]}
                if new_html and (prod["descriptionHtml"] or "").strip() != new_html.strip():
                    product_input["descriptionHtml"] = new_html

                res = gql(token, MUTATION_UPDATE, {
                    "pid": prod["id"],
                    "input": product_input,
                    "variants": [variant_input],
                })
                errs = (res["productUpdate"]["userErrors"] +
                        res["productVariantsBulkUpdate"]["userErrors"])
                if errs:
                    print(f"    ✗ {errs}")
                else:
                    shopify_updated += 1
                    time.sleep(0.2)
            else:
                shopify_updated += 1
        else:
            shopify_ok += 1

    # ── Guardar Excel ──
    if do_excel and excel_changed:
        wb.save(XLSX)
        print(f"\n✓ Excel guardado: {excel_changed} filas actualizadas")

    # ── Resumen ──
    print(f"\n{'='*60}")
    print(f"{'[DRY-RUN] ' if dry else ''}RESUMEN")
    print(f"  Excel  — cambiados: {excel_changed}  ya OK: {excel_ok}")
    if token:
        print(f"  Shopify— {'actualizados' if do_shopify else 'pendientes'}: {shopify_updated}  ya OK: {shopify_ok}  skip(draft): {shopify_skip}")
    if unresolved:
        print(f"\n⚠ Productos que requieren revisión manual ({len(unresolved)}):")
        for u in unresolved:
            print(u)

    if dry:
        print(f"\nPara aplicar: python3 sync_hevea_full.py --apply")
        print(f"Solo Excel:   python3 sync_hevea_full.py --apply --only-excel")
        print(f"Solo Shopify: python3 sync_hevea_full.py --apply --only-shopify")


if __name__ == "__main__":
    main()
