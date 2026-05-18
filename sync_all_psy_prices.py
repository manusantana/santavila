#!/usr/bin/env python3
"""
sync_all_psy_prices.py

Aplica precios psicológicos a TODOS los productos de Shopify, en dos fases:

  1. Productos con SKU en la hoja "20260508 -Todos " del Excel
     → usa col G "Precio Venta Psicologico (con IVA 21%)" como fuente de verdad.

  2. Productos consolidados (SKU empieza por "SV-") no presentes en la hoja
     → aplica psy_price(precio_actual_shopify) directamente.

También actualiza compareAtPrice (×1.30 si price < 50€, ×1.10 si no).

Modos:
  --dry-run (default)  → no toca Shopify, genera psy_prices_report.csv
  --apply              → aplica cambios reales
  --limit N            → procesa solo los primeros N productos de Shopify
  --only-handles a,b   → solo esos handles

Output: psy_prices_report.csv con handle/sku/price_antes/price_despues/
        compare_antes/compare_despues/fuente/status/error
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE  = Path(__file__).resolve().parent
XLSX  = BASE / "Santavila.xlsx"
SHEET = "20260508 -Todos "
REPORT_CSV = BASE / "psy_prices_report.csv"

SHOP = "mueblesexterior.myshopify.com"
API  = f"https://{SHOP}/admin/api/2026-01/graphql.json"

PAUSE_THRESHOLD = 200
PAUSE_SECONDS   = 2.0
UMBRALES = (100, 150, 200, 250, 300, 350, 400, 450, 500, 600, 700, 800, 900,
            1000, 1200, 1500, 1800, 2000)


# ── Redondeo psicológico ─────────────────────────────────────────────────────

def _next_high_ticket(p: float) -> float:
    n = math.ceil(p)
    while n % 10 not in (0, 5, 9):
        n += 1
    return float(n)


def _round_up_to_95(p: float) -> float:
    base = math.floor(p)
    cand = base + 0.95
    if cand + 1e-9 < p:
        cand += 1
    return round(cand, 2)


def _below_umbral(p: float, suffix: float) -> float | None:
    for u in UMBRALES:
        if u <= p <= u * 1.05:
            return round(u - suffix, 2)
    return None


def _round_compare_high(p_compare: float, p_psy: float) -> float:
    lo, hi = p_psy * 1.05, p_psy * 1.12
    for step in (100, 50, 25, 10):
        lo_n = math.ceil(lo / step)
        hi_n = math.floor(hi / step)
        cands = [n * step for n in range(lo_n, hi_n + 1) if n * step > 0]
        if cands:
            return float(min(cands, key=lambda x: abs(x - p_compare)))
    return float(round(p_compare / 10) * 10)


def psy_price(price_bruto: float) -> float:
    if price_bruto < 50:
        return _round_up_to_95(price_bruto)
    if price_bruto <= 500:
        below = _below_umbral(price_bruto, 0.10)
        if below is not None:
            return below
        return _round_up_to_95(price_bruto)
    return _next_high_ticket(price_bruto)


def psy_compare(price_bruto: float) -> float:
    if price_bruto < 50:
        return float(round(price_bruto * 1.30))
    if price_bruto <= 500:
        target = price_bruto * 1.10
        below  = _below_umbral(target, 0.05)
        if below is not None:
            return below
        return _round_up_to_95(target)
    return _round_compare_high(price_bruto * 1.10, psy_price(price_bruto))


def fmt(v: float) -> str:
    return f"{v:.2f}"


def f_eq(a, b, tol=0.01):
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) < tol


# ── Auth ─────────────────────────────────────────────────────────────────────

def read_token() -> str:
    env = (BASE / ".envlocal").read_text(encoding="utf-8")
    m   = re.search(r"^SHOPIFY_ACCESS_TOKEN=(.*)$", env, re.M)
    if not m:
        sys.exit("✗ SHOPIFY_ACCESS_TOKEN no encontrado en .envlocal")
    return m.group(1).strip()


_throttle = {"available": 2000.0}


def gql(token: str, query: str, variables: dict | None = None) -> dict:
    payload  = json.dumps({"query": query, "variables": variables or {}}).encode()
    last_err = None
    for attempt in range(1, 6):
        try:
            req = urllib.request.Request(
                API, data=payload,
                headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read())
            if data.get("errors"):
                raise RuntimeError(json.dumps(data["errors"], ensure_ascii=False))
            ts = data.get("extensions", {}).get("cost", {}).get("throttleStatus", {})
            if ts:
                _throttle["available"] = ts.get("currentlyAvailable", 0)
            if _throttle["available"] < PAUSE_THRESHOLD:
                time.sleep(PAUSE_SECONDS)
            return data["data"]
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")[:400]
            last_err = f"HTTP {e.code}: {body}"
            if e.code == 429:
                time.sleep(float(e.headers.get("Retry-After", 2)))
                continue
        except Exception as e:
            last_err = str(e)
        time.sleep(1.5 * attempt)
    raise RuntimeError(f"GraphQL failed after retries: {last_err}")


# ── Excel: construir mapping handle+sku → precio psicológico ─────────────────

def load_excel_mapping() -> dict[tuple[str, str], float]:
    """Devuelve {(handle, sku): precio_psy} para las filas de la hoja.
    Si un (handle, sku) aparece más de una vez, se excluye (usa psy(shopify))."""
    wb       = openpyxl.load_workbook(XLSX, data_only=True)
    ws       = wb[SHEET]
    seen: dict[tuple[str, str], int] = {}   # key → nº de veces vista
    mapping  = {}
    for r in range(3, ws.max_row + 1):
        handle = ws.cell(r, 2).value
        sku    = ws.cell(r, 3).value
        psy    = ws.cell(r, 7).value  # col G
        if handle and sku and isinstance(psy, (int, float)):
            key = (str(handle).strip(), str(sku).strip())
            seen[key] = seen.get(key, 0) + 1
            if seen[key] == 1:
                mapping[key] = float(psy)
            else:
                mapping.pop(key, None)  # segunda ocurrencia → excluir
    return mapping


# ── Queries y mutaciones GraphQL ─────────────────────────────────────────────

QUERY_ALL_PRODUCTS = """
query($first: Int!, $after: String) {
  products(first: $first, after: $after, query: "status:active") {
    pageInfo { hasNextPage endCursor }
    edges {
      node {
        id handle title
        variants(first: 100) {
          edges {
            node {
              id sku price compareAtPrice
            }
          }
        }
      }
    }
  }
}
"""

MUTATION_BULK_UPDATE = """
mutation($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
  productVariantsBulkUpdate(productId: $productId, variants: $variants) {
    productVariants { id sku price compareAtPrice }
    userErrors { field message }
  }
}
"""


def fetch_all_products(token: str, only_handles: set | None = None,
                       limit: int | None = None) -> list[dict]:
    """Devuelve todos los productos de Shopify con sus variantes."""
    products = []
    cursor   = None
    while True:
        variables = {"first": 50, "after": cursor}
        data      = gql(token, QUERY_ALL_PRODUCTS, variables)
        page      = data["products"]
        for edge in page["edges"]:
            p = edge["node"]
            if only_handles and p["handle"] not in only_handles:
                continue
            products.append(p)
            if limit and len(products) >= limit:
                return products
        if not page["pageInfo"]["hasNextPage"]:
            break
        cursor = page["pageInfo"]["endCursor"]
    return products


# ── Sync principal ────────────────────────────────────────────────────────────

def sync(token, excel_map, *, dry_run=True, limit=None,
         only_handles=None, verbose=True):

    print(f"\n{'━'*80}")
    print(f"{'DRY-RUN' if dry_run else 'APPLY'} — precios psicológicos en todos los productos Shopify")
    print(f"  Excel mapping: {len(excel_map)} SKUs  |  modo: {'preview' if dry_run else 'REAL'}")
    print('━'*80)

    products    = fetch_all_products(token, only_handles=only_handles, limit=limit)
    report_rows = []
    stats = {"productos": len(products), "ok": 0, "sin_cambios": 0,
             "error": 0, "variantes_actualizadas": 0}

    for i, product in enumerate(products, 1):
        product_id = product["id"]
        handle     = product["handle"]
        variants   = [e["node"] for e in product["variants"]["edges"]]

        variants_input     = []
        actions            = []

        for v in variants:
            sku           = (v.get("sku") or "").strip()
            price_actual  = float(v["price"])
            compare_actual = float(v["compareAtPrice"]) if v.get("compareAtPrice") else None

            # Determinar precio psicológico destino y fuente
            key = (handle, sku)
            if key in excel_map:
                price_psy = excel_map[key]
                fuente    = "excel_col_G"
            else:
                # SV-* u otros no en Excel → aplicar psy sobre precio actual
                price_psy = psy_price(price_actual)
                fuente    = "psy(shopify)"

            compare_psy = psy_compare(price_psy)

            cambia_price   = not f_eq(price_actual, price_psy)
            cambia_compare = not f_eq(compare_actual, compare_psy)

            if not cambia_price and not cambia_compare:
                stats["sin_cambios"] += 1
                report_rows.append({
                    "handle": handle, "sku": sku, "fuente": fuente,
                    "price_antes": fmt(price_actual), "price_despues": fmt(price_actual),
                    "compare_antes": fmt(compare_actual) if compare_actual else "",
                    "compare_despues": fmt(compare_actual) if compare_actual else "",
                    "status": "SIN_CAMBIOS", "error": "",
                })
                continue

            inp = {"id": v["id"]}
            if cambia_price:
                inp["price"] = fmt(price_psy)
            if cambia_compare:
                inp["compareAtPrice"] = fmt(compare_psy)
            variants_input.append(inp)
            actions.append({
                "sku": sku, "fuente": fuente,
                "price_antes": price_actual, "price_despues": price_psy if cambia_price else price_actual,
                "compare_antes": compare_actual, "compare_despues": compare_psy if cambia_compare else compare_actual,
                "cambia_price": cambia_price, "cambia_compare": cambia_compare,
            })

        if not variants_input:
            stats["ok"] += 1
            continue

        if verbose:
            print(f"\n[{i}/{len(products)}] {handle}  ({product['title'][:50]})")
            for a in actions:
                cambios = []
                if a["cambia_price"]:
                    cambios.append(f"price {a['price_antes']:.2f}→{a['price_despues']:.2f} [{a['fuente']}]")
                if a["cambia_compare"]:
                    ca = f"{a['compare_antes']:.2f}" if a['compare_antes'] else "—"
                    cambios.append(f"compare {ca}→{a['compare_despues']:.2f}")
                print(f"   • {a['sku'] or '(sin SKU)'}: {' | '.join(cambios)}")

        if dry_run:
            stats["ok"] += 1
            for a in actions:
                stats["variantes_actualizadas"] += 1
                report_rows.append({
                    "handle": handle, "sku": a["sku"], "fuente": a["fuente"],
                    "price_antes": fmt(a["price_antes"]), "price_despues": fmt(a["price_despues"]),
                    "compare_antes": fmt(a["compare_antes"]) if a["compare_antes"] else "",
                    "compare_despues": fmt(a["compare_despues"]) if a["compare_despues"] else "",
                    "status": "DRY_RUN", "error": "",
                })
            continue

        # APPLY
        try:
            mut    = gql(token, MUTATION_BULK_UPDATE,
                         {"productId": product_id, "variants": variants_input})
            errors = mut["productVariantsBulkUpdate"]["userErrors"]
            if errors:
                stats["error"] += 1
                err_str = json.dumps(errors, ensure_ascii=False)[:200]
                if verbose:
                    print(f"   ✗ userErrors: {err_str}")
                for a in actions:
                    stats["variantes_actualizadas"] += 1
                    report_rows.append({
                        "handle": handle, "sku": a["sku"], "fuente": a["fuente"],
                        "price_antes": fmt(a["price_antes"]), "price_despues": fmt(a["price_despues"]),
                        "compare_antes": fmt(a["compare_antes"]) if a["compare_antes"] else "",
                        "compare_despues": fmt(a["compare_despues"]) if a["compare_despues"] else "",
                        "status": "ERROR", "error": err_str,
                    })
            else:
                stats["ok"] += 1
                for a in actions:
                    stats["variantes_actualizadas"] += 1
                    report_rows.append({
                        "handle": handle, "sku": a["sku"], "fuente": a["fuente"],
                        "price_antes": fmt(a["price_antes"]), "price_despues": fmt(a["price_despues"]),
                        "compare_antes": fmt(a["compare_antes"]) if a["compare_antes"] else "",
                        "compare_despues": fmt(a["compare_despues"]) if a["compare_despues"] else "",
                        "status": "OK", "error": "",
                    })
                if verbose:
                    print(f"   ✓ {len(actions)} variantes actualizadas")
        except Exception as e:
            stats["error"] += 1
            if verbose:
                print(f"   ✗ Error: {e}")
            for a in actions:
                report_rows.append({
                    "handle": handle, "sku": a["sku"], "fuente": a["fuente"],
                    "price_antes": fmt(a["price_antes"]), "price_despues": fmt(a["price_despues"]),
                    "compare_antes": "", "compare_despues": "",
                    "status": "ERROR", "error": str(e)[:200],
                })

    # Reporte CSV
    with open(REPORT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "handle", "sku", "fuente",
            "price_antes", "price_despues",
            "compare_antes", "compare_despues",
            "status", "error",
        ])
        w.writeheader()
        w.writerows(report_rows)

    # Resumen
    print(f"\n{'═'*80}")
    print("RESUMEN")
    print('═'*80)
    print(f"  Productos Shopify procesados: {stats['productos']}")
    print(f"    OK / sin errores:           {stats['ok']}")
    print(f"    Con error:                  {stats['error']}")
    print(f"  Variantes {'planificadas' if dry_run else 'actualizadas'}: {stats['variantes_actualizadas']}")
    print(f"  Variantes sin cambios:        {stats['sin_cambios']}")
    print(f"\n  Reporte: {REPORT_CSV.name}")
    if dry_run:
        print(f"\n  ⚠ DRY-RUN — para aplicar: python3 {Path(__file__).name} --apply")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true")
    p.add_argument("--limit", type=int)
    p.add_argument("--only-handles", default=None)
    p.add_argument("--quiet", action="store_true")
    args = p.parse_args()

    token     = read_token()
    excel_map = load_excel_mapping()
    print(f"Excel mapping cargado: {len(excel_map)} SKUs con precio psicológico en col G")

    only = None
    if args.only_handles:
        only = set(h.strip() for h in args.only_handles.split(","))

    sync(token, excel_map,
         dry_run=not args.apply,
         limit=args.limit,
         only_handles=only,
         verbose=not args.quiet)


if __name__ == "__main__":
    main()
