#!/usr/bin/env python3
"""
apply_shipping_categories.py

Clasifica los 281 SKUs de la hoja maestra "20260508 -Todos " en 3 categorías
de envío volumétrico (XS / M / L) y los aplica a Shopify como:
  - metafield  santavila.envio_categoria  (single_line_text_field: "xs" | "m" | "l")
  - tag        envio:xs  /  envio:m  /  envio:l

Reglas de clasificación (basadas en nombre + handle del producto):
  XS  → cojín, funda, limpiador, accesorios pequeños
  M   → mesa auxiliar/centro/lateral, mesa ≤80cm, silla individual, taburete,
        reposapiés, parasol <250cm, accesorios resina pequeños
  L   → mesa comedor, sofá, conjunto, tumbona, banco, balancín, cama balinesa,
        parasol ≥250cm, pérgola, default (cualquier no clasificado arriba)

Tarifas asociadas (se configuran luego en Shopify Shipping Profiles, no aquí):
  XS  → 9,95€ (1ud) / 14,95€ (2) / 19,95€ (3-4) / 24,95€ (5-8) / 29,95€ (9+)
  M   → 29,95€ plano
  L   → 57,95€ plano
  Cualquiera → gratis si subtotal del carrito > 500€

Modos:
  --dry-run (default)  → no toca Shopify, solo genera CSV de clasificación
  --apply              → aplica metafield + tag a cada producto
  --limit N            → procesa solo los primeros N handles (test)
  --only-handles a,b,c → procesa solo handles específicos

Output: shipping_categories_report.csv con la clasificación + estado.

Requisitos:
  - openpyxl instalado (/usr/bin/python3 -m pip install --user openpyxl)
  - .envlocal con SHOPIFY_ACCESS_TOKEN
  - Metafield definition santavila.envio_categoria YA creada en Admin
    (Settings → Custom data → Products → Add definition).
    Tipo: single_line_text_field. Sin esto, el --apply fallará con userError.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE / "Santavila.xlsx"
TODOS_SHEET = "20260508 -Todos "
REPORT_CSV = BASE / "shipping_categories_report.csv"

SHOP = "mueblesexterior.myshopify.com"
API = f"https://{SHOP}/admin/api/2026-01/graphql.json"

PAUSE_THRESHOLD = 200
PAUSE_SECONDS = 2.0


# ── Token ────────────────────────────────────────────────────────────────────

def read_token() -> str:
    for fname in (".env.local", ".envlocal", ".env"):
        p = BASE / fname
        if not p.exists():
            continue
        env = p.read_text(encoding="utf-8")
        m = re.search(r"^SHOPIFY_ACCESS_TOKEN=(.*)$", env, re.M)
        if m:
            return m.group(1).strip().strip('"').strip("'")
    sys.exit("✗ SHOPIFY_ACCESS_TOKEN no encontrado en .env.local / .envlocal / .env")


# ── GraphQL ──────────────────────────────────────────────────────────────────

_throttle = {"available": 2000.0}


def gql(token: str, query: str, variables: dict | None = None) -> dict:
    payload = json.dumps({"query": query, "variables": variables or {}}).encode()
    last_err = None
    for attempt in range(1, 6):
        try:
            req = urllib.request.Request(
                API,
                data=payload,
                headers={
                    "X-Shopify-Access-Token": token,
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read())
            if data.get("errors"):
                raise RuntimeError(json.dumps(data["errors"], ensure_ascii=False))
            ts = data.get("extensions", {}).get("cost", {}).get("throttleStatus", {})
            if ts:
                _throttle["available"] = ts.get("currentlyAvailable", 0)
            if _throttle["available"] < PAUSE_THRESHOLD:
                time.sleep(PAUSE_SECONDS)
            return data["data"]
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")[:500]
            last_err = f"HTTP {e.code}: {body}"
            if e.code == 429:
                wait = float(e.headers.get("Retry-After", 2))
                time.sleep(wait)
                continue
        except (urllib.error.URLError, TimeoutError, RuntimeError) as e:
            last_err = str(e)
        time.sleep(1.5 * attempt)
    raise RuntimeError(f"GraphQL falló tras reintentos: {last_err}")


QUERY_PRODUCT = """
query($handle: String!) {
  productByHandle(handle: $handle) {
    id
    handle
    title
    tags
    metafield(namespace: "santavila", key: "envio_categoria") {
      id
      value
    }
  }
}
"""

MUTATION_UPDATE = """
mutation($input: ProductInput!) {
  productUpdate(input: $input) {
    product {
      id
      tags
      metafield(namespace: "santavila", key: "envio_categoria") { value }
    }
    userErrors { field message }
  }
}
"""


# ── Clasificación ────────────────────────────────────────────────────────────

def categorize(producto: str, handle: str) -> str:
    """Devuelve 'xs' | 'm' | 'l' según el nombre + handle del producto."""
    text = f"{producto or ''} {handle or ''}".lower()

    # XS — cojín, funda, limpiador, accesorios pequeños
    if "cojin" in text or "cojín" in text:
        return "xs"
    if "funda" in text:
        return "xs"
    if "limpiador" in text:
        return "xs"

    # L prioritario — sofá o rinconera (antes de cualquier regla M)
    if "sofa" in text or "sofá" in text or "rinconera" in text:
        return "l"

    # M — silla / taburete / reposapiés
    if "taburete" in text or "reposapi" in text:
        return "m"
    if ("silla" in text or "sillón" in text or "sillon" in text) \
            and "conjunto" not in text and "set" not in text:
        return "m"

    # M — mesa explícitamente pequeña
    if "mesa" in text and any(k in text for k in ("auxiliar", "centro", "baja", "lateral")):
        return "m"

    # M — mesa con medida ≤ 80 cm
    m = re.search(r"(\d+)\s*cm", text)
    if "mesa" in text and m and int(m.group(1)) <= 80:
        return "m"

    # M — accesorio/mobiliario resina pequeño
    if ("accesorio" in text or "mobiliario" in text) and "resina" in text:
        return "m"

    # Parasol: <250cm → M, ≥250cm → L
    if "parasol" in text:
        if m and int(m.group(1)) >= 250:
            return "l"
        return "m" if m else "l"

    # L — todo lo demás (mesa grande, conjunto, tumbona, etc.)
    return "l"


# ── Carga de la hoja ─────────────────────────────────────────────────────────

def load_products() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[TODOS_SHEET]
    rows = []
    seen_handles = set()
    for r in range(3, ws.max_row + 1):
        prov = ws.cell(r, 1).value
        handle = ws.cell(r, 2).value
        sku = ws.cell(r, 3).value
        producto = ws.cell(r, 4).value
        pvp = ws.cell(r, 6).value
        if not handle or not sku:
            continue
        if handle in seen_handles:
            # Varias variantes/filas del mismo producto → solo procesamos el producto 1 vez
            continue
        seen_handles.add(handle)
        rows.append({
            "proveedor": prov,
            "handle": str(handle).strip(),
            "sku": str(sku).strip(),
            "producto": (producto or "").strip(),
            "pvp": float(pvp) if isinstance(pvp, (int, float)) else 0.0,
            "categoria": categorize(producto, handle),
        })
    return rows


# ── Lógica de aplicación ─────────────────────────────────────────────────────

ENVIO_TAGS = {"envio:xs", "envio:m", "envio:l"}


def merge_tags(current: list[str], new_cat: str) -> list[str]:
    """Quita envio:* viejo, añade envio:<new_cat>. Devuelve lista única."""
    cleaned = [t for t in current if t not in ENVIO_TAGS]
    cleaned.append(f"envio:{new_cat}")
    # Únicos preservando orden
    seen = set()
    out = []
    for t in cleaned:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def apply_one(token: str, prod: dict) -> dict:
    """Lee el producto, calcula tags y metafield nuevos, aplica si difieren."""
    data = gql(token, QUERY_PRODUCT, {"handle": prod["handle"]})
    p = data.get("productByHandle")
    if not p:
        return {**prod, "estado": "NO_ENCONTRADO_EN_SHOPIFY"}

    current_tags = p.get("tags") or []
    current_meta = (p.get("metafield") or {}).get("value")
    new_cat = prod["categoria"]
    new_tags = merge_tags(current_tags, new_cat)

    meta_ok = current_meta == new_cat
    tags_ok = sorted(current_tags) == sorted(new_tags)

    if meta_ok and tags_ok:
        return {
            **prod,
            "estado": "SIN_CAMBIOS",
            "tags_antes": ",".join(current_tags),
            "tags_despues": ",".join(new_tags),
            "meta_antes": current_meta or "",
            "meta_despues": new_cat,
        }

    update = gql(token, MUTATION_UPDATE, {
        "input": {
            "id": p["id"],
            "tags": new_tags,
            "metafields": [{
                "namespace": "santavila",
                "key": "envio_categoria",
                "value": new_cat,
                "type": "single_line_text_field",
            }],
        }
    })
    errors = update["productUpdate"]["userErrors"]
    if errors:
        return {
            **prod,
            "estado": "ERROR",
            "tags_antes": ",".join(current_tags),
            "tags_despues": ",".join(new_tags),
            "meta_antes": current_meta or "",
            "meta_despues": new_cat,
            "error": json.dumps(errors, ensure_ascii=False),
        }
    return {
        **prod,
        "estado": "ACTUALIZADO",
        "tags_antes": ",".join(current_tags),
        "tags_despues": ",".join(new_tags),
        "meta_antes": current_meta or "",
        "meta_despues": new_cat,
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="Aplica cambios en Shopify (default: dry-run)")
    parser.add_argument("--limit", type=int, default=None, help="Procesa solo los primeros N handles")
    parser.add_argument("--only-handles", default=None, help="Lista separada por comas")
    args = parser.parse_args()

    dry = not args.apply

    products = load_products()
    if args.only_handles:
        wanted = {h.strip() for h in args.only_handles.split(",") if h.strip()}
        products = [p for p in products if p["handle"] in wanted]
    if args.limit:
        products = products[:args.limit]

    print(f"Productos a procesar: {len(products)}")
    counts = {"xs": 0, "m": 0, "l": 0}
    for p in products:
        counts[p["categoria"]] += 1
    print(f"  XS: {counts['xs']}  ·  M: {counts['m']}  ·  L: {counts['l']}")

    if dry:
        print("\n[DRY-RUN] No se tocará Shopify. Solo se genera CSV con la clasificación.\n")
        with REPORT_CSV.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["proveedor", "handle", "sku", "producto", "pvp", "categoria"])
            for p in products:
                w.writerow([p["proveedor"], p["handle"], p["sku"], p["producto"], f"{p['pvp']:.2f}", p["categoria"]])
        print(f"✓ Reporte escrito: {REPORT_CSV}")
        print("\nPara aplicar realmente: python3 apply_shipping_categories.py --apply")
        return

    token = read_token()
    print("\n[APPLY] Aplicando metafield + tags en Shopify...\n")

    results = []
    for i, p in enumerate(products, 1):
        try:
            res = apply_one(token, p)
        except Exception as e:
            res = {**p, "estado": "EXCEPTION", "error": str(e)}
        results.append(res)
        emoji = {"ACTUALIZADO": "✓", "SIN_CAMBIOS": "·", "ERROR": "✗", "EXCEPTION": "✗", "NO_ENCONTRADO_EN_SHOPIFY": "?"}.get(res["estado"], "?")
        print(f"  {emoji} [{i:3d}/{len(products)}] {p['handle'][:60]:<60} → {p['categoria']}  ({res['estado']})")

    # Reporte
    with REPORT_CSV.open("w", newline="", encoding="utf-8") as f:
        cols = ["proveedor", "handle", "sku", "producto", "pvp", "categoria",
                "estado", "meta_antes", "meta_despues", "tags_antes", "tags_despues", "error"]
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in results:
            w.writerow({c: r.get(c, "") for c in cols})

    # Resumen
    from collections import Counter
    summary = Counter(r["estado"] for r in results)
    print("\n── Resumen ──")
    for k, v in summary.most_common():
        print(f"  {k}: {v}")
    print(f"\n✓ Reporte escrito: {REPORT_CSV}")


if __name__ == "__main__":
    main()
