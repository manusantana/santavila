#!/usr/bin/env python3
"""
fill_psy_column.py

Rellena la columna G "Precio Venta Psicologico (con IVA 21%)" de la hoja
"20260508 -Todos " de Santavila.xlsx aplicando las reglas de redondeo
psicológico a la columna F "Precio Venta (con IVA 21%)".

Reglas (segmentadas por PRICE BRUTO, no por coste):
  < 50€  → termina en .95
  50-500€→ termina en .95; si cae en [umbral, umbral×1.05] → umbral-0.10
  > 500€ → entero sin decimales, terminación 0/5/9 subiendo

Uso:
  python3 fill_psy_column.py           # preview sin guardar
  python3 fill_psy_column.py --apply   # guarda el Excel
"""
from __future__ import annotations

import argparse
import math
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE / "Santavila.xlsx"
SHEET = "20260508 -Todos "
COL_PRICE_BRUTO = 6   # F — Precio Venta (con IVA 21%)
COL_PRICE_PSY   = 7   # G — Precio Venta Psicologico (con IVA 21%)

UMBRALES = (100, 150, 200, 250, 300, 350, 400, 450, 500, 600, 700, 800, 900,
            1000, 1200, 1500, 1800, 2000)


def _next_high_ticket(p: float) -> float:
    n = math.ceil(p)
    while n % 10 not in (0, 5, 9):
        n += 1
    return float(n)


def _round_up_to_95(p: float) -> float:
    base = math.floor(p)
    cand = base + 0.95
    if cand + 1e-9 < p:
        cand += 1
    return round(cand, 2)


def _below_umbral(p: float, suffix: float) -> float | None:
    for u in UMBRALES:
        if u <= p <= u * 1.05:
            return round(u - suffix, 2)
    return None


def psy_price(price_bruto: float) -> float:
    if price_bruto < 50:
        return _round_up_to_95(price_bruto)
    if price_bruto <= 500:
        below = _below_umbral(price_bruto, 0.10)
        if below is not None:
            return below
        return _round_up_to_95(price_bruto)
    return _next_high_ticket(price_bruto)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true", help="Guardar cambios en el Excel")
    args = p.parse_args()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    cambios = []
    sin_cambio = 0

    for r in range(3, ws.max_row + 1):
        bruto = ws.cell(r, COL_PRICE_BRUTO).value
        if not isinstance(bruto, (int, float)):
            continue
        psy = psy_price(float(bruto))
        actual = ws.cell(r, COL_PRICE_PSY).value
        if actual is None or abs(float(actual) - psy) > 0.001:
            cambios.append((r, bruto, psy))
            ws.cell(r, COL_PRICE_PSY).value = psy
        else:
            sin_cambio += 1

    print(f"Filas con precio bruto: {len(cambios) + sin_cambio}")
    print(f"  Que cambian: {len(cambios)}")
    print(f"  Ya correctas: {sin_cambio}")

    if cambios:
        print(f"\nMuestra (primeras 10):")
        print(f"  {'Fila':>5}  {'Bruto':>10}  {'Psicológico':>12}")
        print(f"  {'─'*5}  {'─'*10}  {'─'*12}")
        for row, b, psy in cambios[:10]:
            print(f"  {row:>5}  {b:>10.2f}  {psy:>12.2f}")
        if len(cambios) > 10:
            print(f"  … y {len(cambios) - 10} más")

    if args.apply:
        wb.save(XLSX)
        print(f"\n✅ Guardado en {XLSX.name}")
    else:
        print(f"\n⚠ PREVIEW — usa --apply para guardar en el Excel")


if __name__ == "__main__":
    main()
