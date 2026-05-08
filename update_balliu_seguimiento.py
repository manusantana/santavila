#!/usr/bin/env python3
"""
update_balliu_seguimiento.py

Genera (o regenera) dos hojas en Santavila.xlsx con el seguimiento de tarifas
de Balliu, leyendo todos los PDFs con prefijo YYYYMMDD en proveedores_raw/balliu/.

Semántica de Balliu (distinta a Hevea):
  - Balliu emite DOS tipos de tarifa, ambos sin IVA, con frecuencias distintas:
      • Tarifa CLIENT     → COSTE (lo que paga Santavila a Balliu)
      • Tarifa PVP        → PVP RECOMENDADO (suelo de venta sugerido por Balliu)
  - Cada PDF se clasifica por tipo según contenga "pvp" en el nombre.
  - El "histórico" es una serie temporal POR TIPO (se calculan deltas dentro
    del mismo tipo, no se compara coste con PVP).
  - El "seguimiento" muestra el último valor de cada tipo + márgenes calculados:
      PVP con IVA = PVP sin IVA × 1,21 (lo que paga el cliente final)
      Margen €    = PVP sin IVA − Coste sin IVA
      Margen %    = Margen / PVP sin IVA   (margen bruto, métrica financiera)
      Markup %    = Margen / Coste         (markup sobre coste, lectura pricing)

Identificación: (Producto, Variante, Grupo, Ord). Sin SKU (el PDF no lo trae).
Balliu lista 2 veces algunas combinaciones; el orden de aparición las distingue.

Idempotente. NO toca otras hojas. Snapshots descartados se mueven a
proveedores_raw/balliu/_archived/ (el script no los recoge).

Requiere: pip install openpyxl pdfplumber
"""

import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = Path(__file__).parent
XLSX = BASE / "Santavila.xlsx"
BALLIU_DIR = BASE / "proveedores_raw" / "balliu"

DATE_PREFIX_RE = re.compile(r"^(\d{4})(\d{2})(\d{2})(?:\s|\s*-)")
GROUP_RE = re.compile(r"^G[1-4]$")
PRICE_TOKEN_RE = re.compile(r"^[\d\.]+,\d{2}$")
IVA = 1.21

TIPO_COSTE = "COSTE"
TIPO_PVP = "PVP_RECOMENDADO"

# ── Detección de tipo de tarifa ──────────────────────────────────────────────

def infer_tipo(filename):
    """Heurística: si el nombre del PDF contiene 'pvp' → PVP recomendado.
    Si no → coste cliente. Balliu nombra explícitamente sus PDFs así."""
    return TIPO_PVP if "pvp" in filename.lower() else TIPO_COSTE


# ── Extracción de tarifa desde PDF ───────────────────────────────────────────

def find_dated_pdfs(folder):
    out = []
    skipped = []
    for p in folder.glob("*.pdf"):
        m = DATE_PREFIX_RE.match(p.name)
        if not m:
            continue
        try:
            datetime.strptime(f"{m.group(1)}{m.group(2)}{m.group(3)}", "%Y%m%d")
        except ValueError:
            skipped.append(p.name)
            continue
        date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        out.append((date_str, p))
    if skipped:
        print(f"⚠ PDFs con fecha inválida en el nombre, ignorados: {skipped}", file=sys.stderr)
    return sorted(out)


def group_lines(words, y_tol=3):
    out = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if out and abs(w["top"] - out[-1][0]["top"]) < y_tol:
            out[-1].append(w)
        else:
            out.append([w])
    return out


def parse_line(line_words, x_var=200):
    cols = {"producto": [], "variante": [], "grupo": [], "precio": []}
    for w in line_words:
        t = w["text"]
        if w["x0"] < x_var:
            cols["producto"].append(t)
        elif GROUP_RE.match(t):
            cols["grupo"].append(t)
        elif PRICE_TOKEN_RE.match(t):
            cols["precio"].append(t)
        elif t == "€":
            pass
        else:
            cols["variante"].append(t)
    return {k: " ".join(v).strip() for k, v in cols.items()}


def parse_price(s):
    if not s:
        return None
    s = s.replace("€", "").strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def extract_tarifa(pdf_path):
    """Devuelve lista de dicts {producto, variante, grupo, precio} en orden visual."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for pg_idx, page in enumerate(pdf.pages):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = group_lines(words)
            if not lines:
                continue

            header_idx = None
            for li, ln in enumerate(lines[:8]):
                texts = [w["text"] for w in ln]
                if "Producto" in texts and "Grupo" in texts:
                    header_idx = li
                    break
            if header_idx is None:
                continue

            page_data = []
            page_only = []
            for li, ln in enumerate(lines):
                if li <= header_idx:
                    continue
                top = ln[0]["top"]
                p = parse_line(ln)
                if not p["producto"] and not p["variante"] and not p["grupo"]:
                    continue
                tiene_grupo = bool(GROUP_RE.match(p["grupo"]))
                tiene_precio = bool(p["precio"])
                if tiene_grupo and tiene_precio:
                    page_data.append({
                        "top": top,
                        "producto_inline": p["producto"] or None,
                        "variante": p["variante"],
                        "grupo": p["grupo"],
                        "precio": parse_price(p["precio"]),
                    })
                elif p["producto"] and not tiene_grupo and not tiene_precio:
                    page_only.append({"top": top, "producto": p["producto"]})

            tables = page.extract_tables() or []
            block_sizes = []
            for table in tables:
                current = 0
                for row in table:
                    cell = row[0] if len(row) > 0 else None
                    if cell is not None:
                        if current > 0:
                            block_sizes.append(current)
                        current = 1
                    else:
                        current += 1
                if current > 0:
                    block_sizes.append(current)

            if sum(block_sizes) != len(page_data):
                print(f"⚠ Página {pg_idx+1} de {pdf_path.name}: descuadre bloques({sum(block_sizes)}) vs DATA({len(page_data)})", file=sys.stderr)
                block_sizes = [1] * len(page_data)

            di = 0
            for size in block_sizes:
                block = page_data[di:di + size]
                di += size
                producto = next((d["producto_inline"] for d in block if d["producto_inline"]), None)
                if not producto and block:
                    top_min = min(d["top"] for d in block)
                    top_max = max(d["top"] for d in block)
                    margin = 30
                    candidatos = [o for o in page_only
                                  if top_min - margin <= o["top"] <= top_max + margin]
                    if candidatos:
                        centro = (top_min + top_max) / 2
                        producto = min(candidatos, key=lambda o: abs(o["top"] - centro))["producto"]
                if not producto:
                    producto = "(SIN PRODUCTO)"
                for d in block:
                    rows.append({
                        "producto": producto,
                        "variante": d["variante"],
                        "grupo": d["grupo"],
                        "precio": d["precio"],
                    })
    return rows


def assign_ord(rows):
    seen = defaultdict(int)
    for r in rows:
        k = (r["producto"], r["variante"], r["grupo"])
        seen[k] += 1
        r["ord"] = seen[k]
    return rows


# ── Modelo de datos ──────────────────────────────────────────────────────────

def build_data(snapshots):
    """
    snapshots: lista [(fecha, tipo, rows)] ordenada por fecha asc.
    rows: lista de {producto, variante, grupo, precio, ord}.

    Devuelve dict clave -> {
        producto, variante, grupo, ord,
        series: { TIPO_COSTE: {fecha: precio}, TIPO_PVP: {fecha: precio} }
    }
    """
    products = defaultdict(lambda: {
        "producto": None, "variante": None, "grupo": None, "ord": None,
        "series": {TIPO_COSTE: {}, TIPO_PVP: {}},
    })
    for fecha, tipo, rows in sorted(snapshots, key=lambda x: x[0]):
        for r in rows:
            key = (r["producto"], r["variante"], r["grupo"], r["ord"])
            products[key]["producto"] = r["producto"]
            products[key]["variante"] = r["variante"]
            products[key]["grupo"] = r["grupo"]
            products[key]["ord"] = r["ord"]
            products[key]["series"][tipo][fecha] = r["precio"]
    return products


# ── Estilos ──────────────────────────────────────────────────────────────────

BALLIU_BLUE = "0B5394"
BALLIU_TEAL = "0F6B7A"  # para diferenciar tipo COSTE vs PVP
ROW_A = "DDEEFF"
ROW_B = "EEF5FF"
TOTAL_BG = "F5F5DC"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BLOCKS = "▁▂▃▄▅▆▇█"


def sparkline_text(values):
    nums = [v for v in values if v is not None]
    if len(nums) < 2:
        return ""
    mn, mx = min(nums), max(nums)
    if mx == mn:
        return BLOCKS[3] * len(values)
    rng = mx - mn
    out = ""
    for v in values:
        if v is None:
            out += " "
        else:
            idx = int((v - mn) / rng * (len(BLOCKS) - 1))
            out += BLOCKS[idx]
    return out


def variante_display(v, ord_, has_dups):
    return f"{v} (#{ord_})" if has_dups else v


# ── Hoja "Balliu Histórico" (long, una fila por tipo×fecha) ──────────────────

HEADERS_HIST = [
    "Proveedor", "Producto", "Variante", "Grupo",
    "Tipo tarifa", "Fecha",
    "Valor sin IVA €", "Valor con IVA €",
    "Δ vs anterior €", "Δ vs anterior %",
]
WIDTHS_HIST = [10, 30, 36, 8, 18, 12, 16, 16, 14, 14]


def write_historico(wb, products):
    if "Balliu Histórico" in wb.sheetnames:
        del wb["Balliu Histórico"]
    ws = wb.create_sheet("Balliu Histórico")

    for ci, h in enumerate(HEADERS_HIST, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=BALLIU_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    for ci, w in enumerate(WIDTHS_HIST, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    thick_top = Side(style="medium", color=BALLIU_BLUE)
    border_group_top = Border(left=THIN, right=THIN, top=thick_top, bottom=THIN)

    ord_max = defaultdict(int)
    for k in products:
        prod, var, grp, ord_ = k
        ord_max[(prod, var, grp)] = max(ord_max[(prod, var, grp)], ord_)

    sorted_keys = sorted(products.keys(), key=lambda k: (k[0], k[1], k[2], k[3]))
    rn = 2
    group_idx = 0
    for key in sorted_keys:
        prod, var, grp, ord_ = key
        info = products[key]
        var_disp = variante_display(var, ord_, ord_max[(prod, var, grp)] > 1)
        group_idx += 1
        group_color = ROW_A if group_idx % 2 == 0 else ROW_B
        is_first_subblock = True

        for tipo in (TIPO_COSTE, TIPO_PVP):
            serie = info["series"].get(tipo, {})
            if not serie:
                continue
            prev = None
            fechas_ord = sorted(serie.keys())
            for i_fecha, fecha in enumerate(fechas_ord):
                v = serie[fecha]
                vi = round(v * IVA, 2) if v is not None else None
                d_e = (v - prev) if (v is not None and prev is not None) else None
                d_pct = (d_e / prev * 100) if (d_e is not None and prev) else None
                row_vals = [
                    "Balliu", prod, var_disp, grp,
                    tipo, fecha,
                    v, vi,
                    d_e, d_pct,
                ]
                row_border = border_group_top if (is_first_subblock and i_fecha == 0) else BORDER
                es_primera = (is_first_subblock and i_fecha == 0)
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=rn, column=ci, value=val)
                    c.fill = PatternFill("solid", fgColor=group_color)
                    c.border = row_border
                    if ci in (1, 2, 3, 4) and es_primera:
                        c.font = Font(size=9, bold=True, color=BALLIU_BLUE)
                    else:
                        c.font = Font(size=9)
                    if ci == 5:
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        c.font = Font(size=9, bold=True, color=BALLIU_TEAL if tipo == TIPO_PVP else "1F4E79")
                    elif ci == 6:
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    elif ci in (7, 8):
                        c.number_format = '€ #,##0.00'
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    elif ci == 9:
                        c.number_format = '+€ #,##0.00;-€ #,##0.00;"—"'
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    elif ci == 10:
                        c.number_format = '+0.0"%";-0.0"%";"—"'
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    elif ci == 4:
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")
                rn += 1
                prev = v
            is_first_subblock = False

    last_row = rn - 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_HIST))}{last_row}"
    return last_row - 1


# ── Hoja "Balliu Seguimiento" (wide) ─────────────────────────────────────────

HEADERS_SEG = [
    "Proveedor", "Producto", "Variante", "Grupo", "Estado",
    "Coste sin IVA €", "Última fecha\nCoste",
    "PVP rec.\nsin IVA €", "PVP\ncon IVA €", "Última fecha\nPVP",
    "Margen €\n(sin IVA)",
    "Margen %\nsobre PVP", "Markup %\nsobre coste",
    "Δ Coste %\nvs anterior", "Δ PVP %\nvs anterior",
    "Tendencia\nCoste", "Tendencia\nPVP",
]
WIDTHS_SEG = [10, 30, 34, 7, 16, 13, 13, 13, 13, 13, 13, 12, 13, 13, 13, 12, 12]


def write_seguimiento(wb, snapshots, products):
    if "Balliu Seguimiento" in wb.sheetnames:
        del wb["Balliu Seguimiento"]
    ws = wb.create_sheet("Balliu Seguimiento")

    fechas_coste = sorted({f for f, t, _ in snapshots if t == TIPO_COSTE})
    fechas_pvp = sorted({f for f, t, _ in snapshots if t == TIPO_PVP})
    ult_coste = fechas_coste[-1] if fechas_coste else None
    ult_pvp = fechas_pvp[-1] if fechas_pvp else None

    title = ws.cell(row=1, column=1, value="Balliu — Seguimiento de tarifas")
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=BALLIU_BLUE)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS_SEG))
    ws.row_dimensions[1].height = 30

    meta = (
        f"Última tarifa COSTE: {ult_coste or '—'}  │  "
        f"Última tarifa PVP RECOMENDADO: {ult_pvp or '—'}  │  "
        f"PVP con IVA = PVP sin IVA × 1,21  │  Margen € = PVP − Coste (sin IVA)"
    )
    ws.cell(row=2, column=1, value=meta).font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS_SEG))

    fechas_meta = (
        f"Snapshots COSTE: {' · '.join(fechas_coste) or '(ninguno)'}   "
        f"|   Snapshots PVP: {' · '.join(fechas_pvp) or '(ninguno)'}"
    )
    ws.cell(row=3, column=1, value=fechas_meta).font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(HEADERS_SEG))

    HEADER_ROW = 5
    for ci, h in enumerate(HEADERS_SEG, 1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=BALLIU_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 42

    for ci, w in enumerate(WIDTHS_SEG, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ord_max = defaultdict(int)
    for k in products:
        prod, var, grp, ord_ = k
        ord_max[(prod, var, grp)] = max(ord_max[(prod, var, grp)], ord_)

    sorted_keys = sorted(products.keys(), key=lambda k: (k[0], k[1], k[2], k[3]))
    rn = HEADER_ROW + 1
    for key in sorted_keys:
        prod, var, grp, ord_ = key
        info = products[key]
        var_disp = variante_display(var, ord_, ord_max[(prod, var, grp)] > 1)

        serie_c = info["series"][TIPO_COSTE]
        serie_p = info["series"][TIPO_PVP]
        coste = serie_c[ult_coste] if ult_coste and ult_coste in serie_c else None
        pvp = serie_p[ult_pvp] if ult_pvp and ult_pvp in serie_p else None
        # Si la última fecha global no aplica al producto, coger la última conocida del producto
        last_c = sorted(serie_c.keys())[-1] if serie_c else None
        last_p = sorted(serie_p.keys())[-1] if serie_p else None

        if coste is None and last_c:
            coste = serie_c[last_c]
        if pvp is None and last_p:
            pvp = serie_p[last_p]

        pvp_iva = round(pvp * IVA, 2) if pvp is not None else None
        margen_e = (pvp - coste) if (pvp is not None and coste is not None) else None
        margen_pct = (margen_e / pvp * 100) if (margen_e is not None and pvp) else None
        markup_pct = (margen_e / coste * 100) if (margen_e is not None and coste) else None

        # Estado
        if coste is not None and pvp is not None:
            estado = "COMPLETO"
        elif coste is not None:
            estado = "SOLO COSTE"
        elif pvp is not None:
            estado = "SOLO PVP"
        else:
            estado = "SIN DATOS"

        # Deltas vs anterior dentro del mismo tipo
        def delta_pct(serie):
            fechas = sorted(serie.keys())
            if len(fechas) < 2:
                return None
            p, q = serie[fechas[-2]], serie[fechas[-1]]
            return ((q - p) / p * 100) if (p and q is not None) else None
        d_coste = delta_pct(serie_c)
        d_pvp = delta_pct(serie_p)

        # Tendencias
        tend_c = sparkline_text([serie_c[f] for f in sorted(serie_c.keys())])
        tend_p = sparkline_text([serie_p[f] for f in sorted(serie_p.keys())])

        row_vals = [
            "Balliu", prod, var_disp, grp, estado,
            coste, last_c,
            pvp, pvp_iva, last_p,
            margen_e, margen_pct, markup_pct,
            d_coste, d_pvp,
            tend_c, tend_p,
        ]
        fill = PatternFill("solid", fgColor=ROW_A if rn % 2 == 0 else ROW_B)
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=rn, column=ci, value=val)
            c.fill = fill
            c.border = BORDER
            c.font = Font(size=9)

            if ci in (6, 8, 9, 11):
                c.number_format = '€ #,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (12, 13):
                c.number_format = '0.0"%"'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (14, 15):
                c.number_format = '+0.0"%";-0.0"%";"—"'
                c.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    if val > 0.5:
                        c.font = Font(size=9, color="9C0006", bold=True)
                    elif val < -0.5:
                        c.font = Font(size=9, color="2E7D32", bold=True)
            elif ci in (16, 17):
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Menlo", size=12,
                              color=BALLIU_TEAL if ci == 17 else BALLIU_BLUE)
            elif ci in (7, 10):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if estado == "COMPLETO":
                    c.font = Font(size=9, color="2E7D32", bold=True)
                elif estado in ("SOLO COSTE", "SOLO PVP"):
                    c.font = Font(size=9, color="B45309", bold=True)
                else:
                    c.font = Font(size=9, color="9C0006", italic=True)
            elif ci == 4:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        rn += 1

    last_data = rn - 1

    # Fila TOTAL/MEDIA
    total_row = rn
    label = ws.cell(row=total_row, column=2, value="TOTAL / MEDIA")
    label.font = Font(bold=True, size=10)
    label.fill = PatternFill("solid", fgColor=TOTAL_BG)
    label.alignment = Alignment(horizontal="right", vertical="center")
    label.border = BORDER

    def total_cell(col_idx, formula, fmt):
        c = ws.cell(row=total_row, column=col_idx, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER

    for ci in (6, 8, 9, 11):
        col = get_column_letter(ci)
        total_cell(ci, f"=SUM({col}{HEADER_ROW+1}:{col}{last_data})", '€ #,##0.00')
    for ci in (12, 13):
        col = get_column_letter(ci)
        total_cell(ci, f"=AVERAGE({col}{HEADER_ROW+1}:{col}{last_data})", '0.0"%"')

    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(HEADERS_SEG))}{last_data}"
    ws.freeze_panes = "F6"

    # Color scale sobre Margen % (col 12) — verde si margen alto, rojo si bajo
    margen_col = get_column_letter(12)
    ws.conditional_formatting.add(
        f"{margen_col}{HEADER_ROW+1}:{margen_col}{last_data}",
        ColorScaleRule(
            start_type="num", start_value="0",  start_color="F8696B",
            mid_type="num",   mid_value="35",   mid_color="FFEB84",
            end_type="num",   end_value="55",   end_color="63BE7B",
        )
    )

    return last_data - HEADER_ROW


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not XLSX.exists():
        sys.exit(f"No existe {XLSX}")
    if not BALLIU_DIR.exists():
        sys.exit(f"No existe {BALLIU_DIR}")

    print(f"Leyendo PDFs de {BALLIU_DIR.relative_to(BASE)}...")
    pdfs = find_dated_pdfs(BALLIU_DIR)
    if not pdfs:
        sys.exit("No se encontraron PDFs con prefijo YYYYMMDD")

    snapshots = []
    for fecha, path in pdfs:
        tipo = infer_tipo(path.name)
        rows = extract_tarifa(path)
        rows = assign_ord(rows)
        snapshots.append((fecha, tipo, rows))
        print(f"  {fecha}  [{tipo:18}]  {len(rows):3} filas  ({path.name})")

    products = build_data(snapshots)
    print(f"\n{len(products)} combinaciones únicas (Producto+Variante+Grupo+Ord)")
    cnt_completo = sum(
        1 for p in products.values()
        if p["series"][TIPO_COSTE] and p["series"][TIPO_PVP]
    )
    cnt_solo_coste = sum(
        1 for p in products.values()
        if p["series"][TIPO_COSTE] and not p["series"][TIPO_PVP]
    )
    cnt_solo_pvp = sum(
        1 for p in products.values()
        if not p["series"][TIPO_COSTE] and p["series"][TIPO_PVP]
    )
    print(f"  con coste y PVP:  {cnt_completo}")
    print(f"  solo coste:        {cnt_solo_coste}")
    print(f"  solo PVP:          {cnt_solo_pvp}")

    print(f"\nAbriendo {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX)
    print(f"  Hojas existentes: {wb.sheetnames}")

    n_hist = write_historico(wb, products)
    print(f"\n✓ 'Balliu Histórico' regenerado — {n_hist} filas")

    n_seg = write_seguimiento(wb, snapshots, products)
    print(f"✓ 'Balliu Seguimiento' regenerado — {n_seg} filas")

    # Limpiar DefinedNames internos duplicados que openpyxl arrastra al recargar
    # (autofiltros sin localSheetId que rompen el archivo en Excel)
    for n in [x for x in wb.defined_names if x.startswith("_xlnm.")]:
        del wb.defined_names[n]

    wb.save(XLSX)
    print(f"\n✅ {XLSX.name} actualizado")
    print(f"   Hojas finales: {wb.sheetnames}")


if __name__ == "__main__":
    main()
