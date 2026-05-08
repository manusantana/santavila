#!/usr/bin/env python3
"""
sync_prices_to_shopify.py

Sincroniza precios y costes de la hoja "20260508 -Todos " hacia Shopify.

Lee:
  - Handle Shopify (col B), SKU (col C)
  - Coste sin IVA (col E) → inventoryItem.cost
  - Precio Venta con IVA 21% (col F) → variant.price

Cruza por (Handle, SKU) y hace bulk update por producto.

Modos:
  --dry-run (default)  → no toca Shopify, solo imprime y genera CSV de cambios
  --apply              → aplica los cambios reales
  --limit N            → procesa solo los primeros N productos (útil para test)
  --only-handles a,b,c → procesa solo handles específicos
  --skip-cost          → no actualiza el coste (solo el price)
  --skip-price         → no actualiza el price (solo el coste)

Output: sync_prices_report.csv con (handle, sku, accion, precio_antes, precio_despues,
       coste_antes, coste_despues, status, error).

Uso:
  python3 sync_prices_to_shopify.py                    # dry-run completo
  python3 sync_prices_to_shopify.py --limit 5          # dry-run de 5 productos
  python3 sync_prices_to_shopify.py --apply --limit 1  # aplicar a 1 producto (test real)
  python3 sync_prices_to_shopify.py --apply            # aplicar todo
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE / "Santavila.xlsx"
TODOS_SHEET = "20260508 -Todos "
REPORT_CSV = BASE / "sync_prices_report.csv"

SHOP = "mueblesexterior.myshopify.com"
API = f"https://{SHOP}/admin/api/2026-01/graphql.json"

# Bucket: 2000 pts, restore 100/s. Cada query ~10 pts, cada mutación ~10-15 pts.
# Pausa preventiva si bajamos de PAUSE_THRESHOLD pts.
PAUSE_THRESHOLD = 200
PAUSE_SECONDS = 2.0


def read_token() -> str:
    env = (BASE / ".envlocal").read_text(encoding="utf-8")
    m = re.search(r"^SHOPIFY_ACCESS_TOKEN=(.*)$", env, re.M)
    if not m:
        sys.exit("✗ SHOPIFY_ACCESS_TOKEN no encontrado en .envlocal")
    return m.group(1).strip()


_throttle_state = {"available": 2000.0, "restore": 100.0}


def gql(token: str, query: str, variables: dict | None = None) -> dict:
    payload = json.dumps({"query": query, "variables": variables or {}}).encode()
    last_err = None
    for attempt in range(1, 6):
        try:
            req = urllib.request.Request(
                API,
                data=payload,
                headers={
                    "X-Shopify-Access-Token": token,
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read())
            if data.get("errors"):
                raise RuntimeError(json.dumps(data["errors"], ensure_ascii=False))
            ext = data.get("extensions", {})
            cost = ext.get("cost", {})
            ts = cost.get("throttleStatus", {})
            if ts:
                _throttle_state["available"] = ts.get("currentlyAvailable", 0)
                _throttle_state["restore"] = ts.get("restoreRate", 100)
            # Pausa preventiva
            if _throttle_state["available"] < PAUSE_THRESHOLD:
                time.sleep(PAUSE_SECONDS)
            return data["data"]
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")[:500]
            last_err = f"HTTP {e.code}: {body}"
            if e.code == 429:
                wait = float(e.headers.get("Retry-After", 2))
                time.sleep(wait)
                continue
        except (urllib.error.URLError, TimeoutError, RuntimeError) as e:
            last_err = str(e)
        time.sleep(1.5 * attempt)
    raise RuntimeError(f"GraphQL request failed after retries: {last_err}")


# ── Cargar hoja ──────────────────────────────────────────────────────────────

def load_todos():
    """Devuelve dict {handle: [{sku, precio_iva, coste, fila, producto, proveedor}]}."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[TODOS_SHEET]
    by_handle = defaultdict(list)
    for r in range(3, ws.max_row + 1):
        prov = ws.cell(r, 1).value
        handle = ws.cell(r, 2).value
        sku = ws.cell(r, 3).value
        prod = ws.cell(r, 4).value
        coste = ws.cell(r, 5).value
        precio_iva = ws.cell(r, 6).value
        if not handle or not sku:
            continue
        if not isinstance(coste, (int, float)) or not isinstance(precio_iva, (int, float)):
            continue
        by_handle[handle].append({
            "sku": str(sku).strip(),
            "precio_iva": float(precio_iva),
            "coste": float(coste),
            "fila": r,
            "producto": prod,
            "proveedor": prov,
        })
    return by_handle


# ── GraphQL ──────────────────────────────────────────────────────────────────

QUERY_PRODUCT_BY_HANDLE = """
query($handle: String!) {
  productByHandle(handle: $handle) {
    id
    handle
    title
    variants(first: 50) {
      edges {
        node {
          id
          sku
          price
          inventoryItem {
            id
            unitCost { amount }
          }
        }
      }
    }
  }
}
"""

MUTATION_BULK_UPDATE = """
mutation($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
  productVariantsBulkUpdate(productId: $productId, variants: $variants) {
    productVariants {
      id
      sku
      price
      inventoryItem { id unitCost { amount } }
    }
    userErrors { field message }
  }
}
"""


def fmt_price(v: float) -> str:
    """Shopify quiere strings como "1185.80" (sin separadores de miles, dos decimales)."""
    return f"{v:.2f}"


def f_eq(a, b, tol=0.01):
    """Comparación de floats con tolerancia."""
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) < tol


# ── Sincronización ───────────────────────────────────────────────────────────

def sync(token, by_handle, *, dry_run=True, limit=None,
         only_handles=None, skip_cost=False, skip_price=False, verbose=True):
    handles = list(by_handle.keys())
    if only_handles:
        handles = [h for h in handles if h in set(only_handles)]
    if limit:
        handles = handles[:limit]

    report_rows = []
    stats = {
        "handles_total": len(handles),
        "handles_ok": 0,
        "handles_no_existe": 0,
        "handles_error": 0,
        "variants_actualizadas": 0,
        "variants_sin_match_sku": 0,
        "variants_sin_cambios": 0,
        "variants_error": 0,
    }

    print(f"\n{'━' * 80}")
    print(f"{'DRY-RUN' if dry_run else 'APPLY'} | {len(handles)} handles | "
          f"price={'no' if skip_price else 'sí'} | cost={'no' if skip_cost else 'sí'}")
    print('━' * 80)

    for i, handle in enumerate(handles, 1):
        rows = by_handle[handle]
        try:
            data = gql(token, QUERY_PRODUCT_BY_HANDLE, {"handle": handle})
        except Exception as e:
            stats["handles_error"] += 1
            if verbose:
                print(f"\n[{i}/{len(handles)}] {handle}  ✗ Error consultando: {e}")
            for r in rows:
                report_rows.append({
                    "handle": handle, "sku": r["sku"], "accion": "ERROR_QUERY",
                    "precio_antes": "", "precio_despues": "",
                    "coste_antes": "", "coste_despues": "",
                    "status": "ERROR", "error": str(e)[:200],
                })
            continue

        product = data.get("productByHandle")
        if not product:
            stats["handles_no_existe"] += 1
            if verbose:
                print(f"\n[{i}/{len(handles)}] {handle}  ⚠ NO existe en Shopify")
            for r in rows:
                report_rows.append({
                    "handle": handle, "sku": r["sku"], "accion": "NO_EXISTE",
                    "precio_antes": "", "precio_despues": fmt_price(r["precio_iva"]),
                    "coste_antes": "", "coste_despues": fmt_price(r["coste"]),
                    "status": "WARN", "error": "Handle no existe en Shopify",
                })
            continue

        product_id = product["id"]
        variants_shopify = {
            v["node"]["sku"]: v["node"]
            for v in product["variants"]["edges"]
            if v["node"].get("sku")
        }

        # Detectar SKUs duplicados dentro del mismo handle (caso Hevea: 557-010884,
        # 557-010147, 557-1563 — el proveedor reusa SKU para productos distintos).
        # Resolución: mantener solo la fila cuyo coste es más cercano al coste
        # ACTUAL de Shopify. Es la única señal fiable de qué producto está
        # realmente vivo en la tienda.
        from collections import Counter
        sku_counts = Counter(r["sku"] for r in rows)
        dup_skus = {sku for sku, n in sku_counts.items() if n > 1}
        if dup_skus:
            # Por cada SKU duplicado, elegir la fila más cercana al coste de Shopify
            kept_rows = []
            descartados = []
            for sku in sku_counts:
                grupo = [r for r in rows if r["sku"] == sku]
                if len(grupo) == 1:
                    kept_rows.extend(grupo)
                    continue
                v = variants_shopify.get(sku)
                if not v or not v["inventoryItem"].get("unitCost"):
                    # No hay coste actual en Shopify → cogemos la primera y avisamos
                    kept_rows.append(grupo[0])
                    for desc in grupo[1:]:
                        descartados.append((desc, "duplicado sin ground truth en Shopify"))
                    continue
                coste_shopify = float(v["inventoryItem"]["unitCost"]["amount"])
                # Ordenar por proximidad de coste; el más cercano gana
                grupo.sort(key=lambda r: abs(r["coste"] - coste_shopify))
                kept_rows.append(grupo[0])
                for desc in grupo[1:]:
                    descartados.append((desc, f"otra fila más cercana al coste actual {coste_shopify}"))

            if verbose:
                print(f"\n[{i}/{len(handles)}] {handle}  ⚠ SKU duplicado {sorted(dup_skus)} → elegida fila por coste cercano")
                for r in kept_rows:
                    if r["sku"] in dup_skus:
                        print(f"   ✓ R{r['fila']}: {r['producto']!r:60} coste={r['coste']}")
                for desc, motivo in descartados:
                    print(f"   ✗ R{desc['fila']}: {desc['producto']!r:60} (descartado: {motivo})")
            for desc, motivo in descartados:
                report_rows.append({
                    "handle": handle, "sku": desc["sku"], "accion": "DUPLICADO_DESCARTADO",
                    "precio_antes": "", "precio_despues": fmt_price(desc["precio_iva"]),
                    "coste_antes": "", "coste_despues": fmt_price(desc["coste"]),
                    "status": "WARN",
                    "error": f"R{desc['fila']} '{desc['producto']}' descartada: {motivo}",
                })
            rows = kept_rows  # usar solo las filas elegidas

        # Construir variants_input
        variants_input = []
        actions_per_variant = []
        for r in rows:
            v = variants_shopify.get(r["sku"])
            if not v:
                stats["variants_sin_match_sku"] += 1
                report_rows.append({
                    "handle": handle, "sku": r["sku"], "accion": "SKU_NO_EN_SHOPIFY",
                    "precio_antes": "", "precio_despues": fmt_price(r["precio_iva"]),
                    "coste_antes": "", "coste_despues": fmt_price(r["coste"]),
                    "status": "WARN",
                    "error": f"SKU {r['sku']} no encontrado en variantes del producto",
                })
                if verbose:
                    print(f"  ⚠ SKU {r['sku']} no en producto {handle}")
                continue

            precio_actual = float(v["price"])
            coste_actual = (
                float(v["inventoryItem"]["unitCost"]["amount"])
                if v["inventoryItem"].get("unitCost") else None
            )
            precio_nuevo = r["precio_iva"]
            coste_nuevo = r["coste"]

            cambia_precio = (not skip_price) and not f_eq(precio_actual, precio_nuevo)
            cambia_coste = (not skip_cost) and not f_eq(coste_actual, coste_nuevo)

            if not cambia_precio and not cambia_coste:
                stats["variants_sin_cambios"] += 1
                report_rows.append({
                    "handle": handle, "sku": r["sku"], "accion": "SIN_CAMBIOS",
                    "precio_antes": fmt_price(precio_actual),
                    "precio_despues": fmt_price(precio_actual),
                    "coste_antes": fmt_price(coste_actual) if coste_actual is not None else "",
                    "coste_despues": fmt_price(coste_actual) if coste_actual is not None else "",
                    "status": "OK", "error": "",
                })
                continue

            inp = {"id": v["id"]}
            if cambia_precio:
                inp["price"] = fmt_price(precio_nuevo)
            if cambia_coste:
                inp["inventoryItem"] = {"cost": fmt_price(coste_nuevo)}
            variants_input.append(inp)
            actions_per_variant.append({
                "sku": r["sku"],
                "precio_actual": precio_actual,
                "precio_nuevo": precio_nuevo if cambia_precio else precio_actual,
                "coste_actual": coste_actual,
                "coste_nuevo": coste_nuevo if cambia_coste else coste_actual,
                "cambia_precio": cambia_precio,
                "cambia_coste": cambia_coste,
            })

        if not variants_input:
            stats["handles_ok"] += 1
            if verbose:
                print(f"[{i}/{len(handles)}] {handle}  ─ sin cambios necesarios")
            continue

        if verbose:
            print(f"\n[{i}/{len(handles)}] {handle}  ({product['title'][:50]})")
            for a in actions_per_variant:
                cambios = []
                if a["cambia_precio"]:
                    cambios.append(f"price {a['precio_actual']:.2f}→{a['precio_nuevo']:.2f}")
                if a["cambia_coste"]:
                    co_a = f"{a['coste_actual']:.2f}" if a['coste_actual'] is not None else "—"
                    cambios.append(f"cost {co_a}→{a['coste_nuevo']:.2f}")
                print(f"   • {a['sku']}: {' | '.join(cambios)}")

        if dry_run:
            stats["handles_ok"] += 1
            for a in actions_per_variant:
                stats["variants_actualizadas"] += 1
                report_rows.append({
                    "handle": handle, "sku": a["sku"], "accion": "DRY_RUN",
                    "precio_antes": fmt_price(a["precio_actual"]),
                    "precio_despues": fmt_price(a["precio_nuevo"]),
                    "coste_antes": fmt_price(a["coste_actual"]) if a["coste_actual"] is not None else "",
                    "coste_despues": fmt_price(a["coste_nuevo"]),
                    "status": "PLAN", "error": "",
                })
            continue

        # APPLY: ejecutar mutación
        try:
            mut = gql(token, MUTATION_BULK_UPDATE,
                      {"productId": product_id, "variants": variants_input})
            errors = mut["productVariantsBulkUpdate"]["userErrors"]
            if errors:
                stats["handles_error"] += 1
                err_str = json.dumps(errors, ensure_ascii=False)[:200]
                if verbose:
                    print(f"   ✗ userErrors: {err_str}")
                for a in actions_per_variant:
                    stats["variants_error"] += 1
                    report_rows.append({
                        "handle": handle, "sku": a["sku"], "accion": "ERROR_MUTATION",
                        "precio_antes": fmt_price(a["precio_actual"]),
                        "precio_despues": fmt_price(a["precio_nuevo"]),
                        "coste_antes": fmt_price(a["coste_actual"]) if a["coste_actual"] is not None else "",
                        "coste_despues": fmt_price(a["coste_nuevo"]),
                        "status": "ERROR", "error": err_str,
                    })
            else:
                stats["handles_ok"] += 1
                for a in actions_per_variant:
                    stats["variants_actualizadas"] += 1
                    report_rows.append({
                        "handle": handle, "sku": a["sku"], "accion": "ACTUALIZADO",
                        "precio_antes": fmt_price(a["precio_actual"]),
                        "precio_despues": fmt_price(a["precio_nuevo"]),
                        "coste_antes": fmt_price(a["coste_actual"]) if a["coste_actual"] is not None else "",
                        "coste_despues": fmt_price(a["coste_nuevo"]),
                        "status": "OK", "error": "",
                    })
                if verbose:
                    print(f"   ✓ {len(actions_per_variant)} variantes actualizadas")
        except Exception as e:
            stats["handles_error"] += 1
            if verbose:
                print(f"   ✗ Error mutación: {e}")
            for a in actions_per_variant:
                stats["variants_error"] += 1
                report_rows.append({
                    "handle": handle, "sku": a["sku"], "accion": "ERROR_MUTATION",
                    "precio_antes": "", "precio_despues": fmt_price(a["precio_nuevo"]),
                    "coste_antes": "", "coste_despues": fmt_price(a["coste_nuevo"]),
                    "status": "ERROR", "error": str(e)[:200],
                })

    # Escribir reporte
    with open(REPORT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "handle", "sku", "accion",
            "precio_antes", "precio_despues",
            "coste_antes", "coste_despues",
            "status", "error",
        ])
        w.writeheader()
        w.writerows(report_rows)

    return stats


def print_stats(stats, dry_run):
    print("\n" + "═" * 80)
    print("RESUMEN")
    print("═" * 80)
    print(f"  Handles totales:                {stats['handles_total']}")
    print(f"    OK:                           {stats['handles_ok']}")
    print(f"    No existen en Shopify:        {stats['handles_no_existe']}")
    print(f"    Con error:                    {stats['handles_error']}")
    print(f"  Variantes:")
    print(f"    {'Planificadas' if dry_run else 'Actualizadas'}: {stats['variants_actualizadas']}")
    print(f"    Sin cambios necesarios:       {stats['variants_sin_cambios']}")
    print(f"    SKU no en Shopify:            {stats['variants_sin_match_sku']}")
    print(f"    Con error:                    {stats['variants_error']}")
    print(f"\n  Reporte CSV: {REPORT_CSV.relative_to(BASE)}")
    if dry_run:
        print(f"\n  ⚠ MODO DRY-RUN: no se ha tocado Shopify.")
        print(f"  Para aplicar: python3 sync_prices_to_shopify.py --apply")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true", help="Aplicar cambios reales (default: dry-run)")
    p.add_argument("--limit", type=int, default=None, help="Procesar solo N handles")
    p.add_argument("--only-handles", default=None, help="Lista CSV de handles a procesar")
    p.add_argument("--skip-cost", action="store_true", help="No actualizar coste")
    p.add_argument("--skip-price", action="store_true", help="No actualizar precio")
    p.add_argument("--quiet", action="store_true", help="Menos verbosidad")
    args = p.parse_args()

    token = read_token()
    by_handle = load_todos()
    print(f"Cargadas {sum(len(v) for v in by_handle.values())} filas en "
          f"{len(by_handle)} handles únicos desde '{TODOS_SHEET}'")

    only = None
    if args.only_handles:
        only = [h.strip() for h in args.only_handles.split(",") if h.strip()]

    stats = sync(
        token, by_handle,
        dry_run=not args.apply,
        limit=args.limit,
        only_handles=only,
        skip_cost=args.skip_cost,
        skip_price=args.skip_price,
        verbose=not args.quiet,
    )
    print_stats(stats, dry_run=not args.apply)


if __name__ == "__main__":
    main()
